from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Form, Request, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
import csv
import io
import hashlib
import json
import httpx
import logging
import re
import calendar

from app.db.session import get_db_session
from app.core.rate_limit import enforce_rate_limit
from app.models.upload import Upload
from app.models.job import Job
from app.models.activity import Activity, ActivityType
from app.models.budget import BudgetTarget, KpiTarget
from app.models.company import Company
from app.models.contact import Contact
from app.models.deal import Deal
from app.models.content_item import ContentItem, ContentItemStatus
from app.models.content_task import ContentTask, ContentTaskStatus, ContentTaskPriority
from app.models.calendar import CalendarEntry
from app.models.user import User, UserRole
from app.models.user_category import UserCategory
from app.models.upload_audit_log import UploadAuditLog
from app.api.deps import get_current_user, get_org_id, is_demo_user, require_writable_user
from app.core.config import get_settings

router = APIRouter(prefix="/uploads", tags=["uploads"]) 

logger = logging.getLogger(__name__)


def _safe_add_upload_audit_log(db: Session, **payload: Any) -> None:
    """
    Audit logging must never block the actual user action.
    Use a savepoint so broken/misaligned audit constraints in production
    do not abort the surrounding delete/import transaction.
    """
    try:
        with db.begin_nested():
            db.add(UploadAuditLog(**payload))
            db.flush()
    except Exception:
        try:
            logger.warning("upload audit log skipped", exc_info=True)
        except Exception:
            pass


@router.delete("/{upload_id}")
def delete_upload(
    upload_id: int,
    db: Session = Depends(get_db_session),
    current_user: User = Depends(require_writable_user),
) -> Dict[str, Any]:
    """
    Delete an upload AND everything imported from it (within the same organization).
    This is critical for "try import -> delete -> platform clean" workflows.
    """
    org = get_org_id(current_user)
    can_manage_all = current_user.role in {UserRole.admin, UserRole.editor}

    upload = db.query(Upload).filter(Upload.id == int(upload_id), Upload.organization_id == org).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    if not can_manage_all and int(getattr(upload, "owner_id", 0) or 0) != int(current_user.id):
        raise HTTPException(status_code=403, detail="Not allowed")

    deleted: Dict[str, int] = {}

    def _add_deleted(name: str, n: int):
        if not n:
            return
        deleted[name] = int(deleted.get(name, 0) or 0) + int(n)

    def _del(model, name: str):
        try:
            q = db.query(model).filter(model.organization_id == org, model.source_upload_id == int(upload_id))  # type: ignore[attr-defined]
            n = q.delete(synchronize_session=False)
            _add_deleted(name, int(n or 0))
        except Exception:
            pass

    # Domain objects created by imports / smart-import.
    _del(CalendarEntry, "calendar_entries")
    _del(Activity, "activities")
    _del(ContentTask, "content_tasks")
    _del(ContentItem, "content_items")
    _del(BudgetTarget, "budget_targets")
    _del(KpiTarget, "kpi_targets")
    _del(Deal, "deals")
    _del(Contact, "contacts")
    _del(Company, "companies")
    _del(UserCategory, "user_categories")

    # --- Legacy cleanup (for imports done before provenance tagging existed) ---
    # If a previous deployment created records without source_upload_id, deleting the upload would
    # otherwise remove only the upload row, leaving imported objects behind.
    try:
        from datetime import timedelta, date as _date

        name_lower = str(getattr(upload, "original_name", "") or "").lower()
        is_tabular = name_lower.endswith(".csv") or name_lower.endswith(".xlsx") or name_lower.endswith(".xls")
        owner_id = int(getattr(upload, "owner_id", 0) or 0)
        t0 = getattr(upload, "created_at", None)

        # Only attempt legacy cleanup for tabular uploads; keep it conservative for non-import files.
        if is_tabular and t0 and owner_id > 0:
            window_start = t0 - timedelta(hours=4)
            window_end = t0 + timedelta(hours=4)

            # 1) Try high-confidence cleanup: match activities by signatures derived from the file content.
            sig3: set[tuple[str, Optional[_date], Optional[_date]]] = set()
            sig4: set[tuple[str, Optional[_date], Optional[_date], str]] = set()
            min_d: Optional[_date] = None
            max_d: Optional[_date] = None

            try:
                content_bytes = bytes(getattr(upload, "content", None) or b"")
                headers: List[str] = []
                rows: List[Dict[str, Any]] = []

                if content_bytes:
                    if name_lower.endswith(".csv"):
                        text = content_bytes.decode("utf-8", errors="ignore")
                        reader = csv.DictReader(io.StringIO(text))
                        headers = list(reader.fieldnames or [])
                        for i, row in enumerate(reader):
                            rows.append(row)
                            if i >= 4999:
                                break
                    else:
                        tables = _extract_xlsx_tables(content_bytes)
                        picked = None
                        for t in tables:
                            if str(t.get("sheet") or "").strip().lower() == "mediaplan":
                                picked = t
                                break
                        if not picked and tables:
                            picked = max(tables, key=lambda t: len(t.get("rows") or []))
                        headers = list((picked or {}).get("headers") or [])
                        rows = list((picked or {}).get("rows") or [])

                if headers and rows:
                    m = _suggest_mapping_activities(headers)
                    h_title = m.get("title") or ("title" if "title" in {str(h or "").strip().lower() for h in headers} else None)
                    h_start = m.get("start") or ("start" if "start" in {str(h or "").strip().lower() for h in headers} else None)
                    h_end = m.get("end") or ("end" if "end" in {str(h or "").strip().lower() for h in headers} else None)
                    h_cat = m.get("category") or ("category" if "category" in {str(h or "").strip().lower() for h in headers} else None)

                    def _bucket_for_mediaplan_row(row: Dict[str, Any]) -> Optional[str]:
                        try:
                            section = str(row.get("section") or "").strip()
                            medium = str(row.get("medium") or "").strip()
                            form = str(row.get("form") or "").strip()
                            if not (section or medium or form):
                                return None
                            blob = f"{section} {medium} {form}".lower()
                            if "sap intern" in section.lower():
                                return "SAP intern"
                            if "paid social" in blob:
                                return "Paid Social"
                            if "organic" in blob:
                                return "Organic Social"
                            if " print" in blob or blob.endswith("print") or "magazin gedruckte" in blob:
                                return "Print"
                            return "Online Ads"
                        except Exception:
                            return None

                    for r in rows[:5000]:
                        try:
                            raw_title = r.get(h_title) if h_title else r.get("title")
                            title = str(raw_title or "").strip()
                            if not title:
                                continue

                            bucket = _bucket_for_mediaplan_row(r)
                            raw_cat = r.get(h_cat) if h_cat else r.get("category")
                            cat = str(raw_cat or "").strip()
                            cat_eff = bucket or cat

                            sd_raw = r.get(h_start) if h_start else r.get("start")
                            ed_raw = r.get(h_end) if h_end else r.get("end")
                            sd_dt = _parse_datetime_loose(sd_raw)
                            ed_dt = _parse_datetime_loose(ed_raw)
                            sd = sd_dt.date() if sd_dt else None
                            ed = ed_dt.date() if ed_dt else None

                            tl = title.lower()
                            sig3.add((tl, sd, ed))
                            if cat_eff:
                                sig4.add((tl, sd, ed, str(cat_eff).strip().lower()))

                            for d in (sd, ed):
                                if not d:
                                    continue
                                min_d = d if min_d is None else min(min_d, d)
                                max_d = d if max_d is None else max(max_d, d)
                        except Exception:
                            continue
            except Exception:
                pass

            legacy_activity_ids: List[int] = []
            try:
                if sig3:
                    q = (
                        db.query(Activity)
                        .filter(
                            Activity.organization_id == org,
                            Activity.owner_id == owner_id,
                            Activity.source_upload_id.is_(None),
                        )
                    )
                    # Narrow down candidates by date range when possible.
                    if min_d and max_d:
                        pad = timedelta(days=2)
                        q = q.filter(
                            Activity.start_date.isnot(None),
                            Activity.start_date >= (min_d - pad),
                            Activity.start_date <= (max_d + pad),
                        )
                    candidates = q.all()
                    for a in candidates:
                        try:
                            tl = str(a.title or "").strip().lower()
                            sd = getattr(a, "start_date", None)
                            ed = getattr(a, "end_date", None)
                            key3 = (tl, sd, ed)
                            if key3 in sig3:
                                legacy_activity_ids.append(int(a.id))
                                continue
                            cat_l = str(getattr(a, "category_name", "") or "").strip().lower()
                            if cat_l and (tl, sd, ed, cat_l) in sig4:
                                legacy_activity_ids.append(int(a.id))
                        except Exception:
                            continue
            except Exception:
                legacy_activity_ids = []

            if legacy_activity_ids:
                # Calendar entries linked to these activities
                try:
                    n = (
                        db.query(CalendarEntry)
                        .filter(
                            CalendarEntry.organization_id == org,
                            CalendarEntry.activity_id.in_(legacy_activity_ids),
                        )
                        .delete(synchronize_session=False)
                    )
                    _add_deleted("calendar_entries", int(n or 0))
                except Exception:
                    pass
                try:
                    n = (
                        db.query(Activity)
                        .filter(Activity.organization_id == org, Activity.id.in_(legacy_activity_ids))
                        .delete(synchronize_session=False)
                    )
                    _add_deleted("activities", int(n or 0))
                except Exception:
                    pass

            # 2) Time-window cleanup for other legacy objects created by this import
            # (only those still missing provenance).
            def _del_legacy_time(model, name: str, *, time_field: str, extra_filters: Optional[List[Any]] = None):
                try:
                    col = getattr(model, time_field)
                    q = db.query(model).filter(
                        model.organization_id == org,  # type: ignore[attr-defined]
                        model.source_upload_id.is_(None),  # type: ignore[attr-defined]
                        col >= window_start,
                        col <= window_end,
                    )
                    if extra_filters:
                        for f in extra_filters:
                            q = q.filter(f)
                    n = q.delete(synchronize_session=False)
                    _add_deleted(name, int(n or 0))
                except Exception:
                    pass

            # Calendar entries created by imports (activity/content) for this owner in the same window
            _del_legacy_time(CalendarEntry, "calendar_entries", time_field="created_at", extra_filters=[CalendarEntry.owner_id == owner_id])
            # Content created by smart import
            _del_legacy_time(ContentTask, "content_tasks", time_field="created_at", extra_filters=[ContentTask.owner_id == owner_id])
            _del_legacy_time(ContentItem, "content_items", time_field="created_at", extra_filters=[ContentItem.owner_id == owner_id])
            # Budget/KPI targets may be upserted; prefer updated_at for cleanup.
            _del_legacy_time(BudgetTarget, "budget_targets", time_field="updated_at")
            _del_legacy_time(KpiTarget, "kpi_targets", time_field="updated_at")
            # CRM entities created via import (no owner_id field; use window + org + missing provenance)
            _del_legacy_time(Deal, "deals", time_field="created_at")
            _del_legacy_time(Contact, "contacts", time_field="created_at")
            _del_legacy_time(Company, "companies", time_field="created_at")
            # User categories are user-scoped; delete only for the owner.
            _del_legacy_time(UserCategory, "user_categories", time_field="created_at", extra_filters=[UserCategory.user_id == owner_id])
    except Exception:
        pass

    # Jobs linked to this upload (best-effort: rq_id contains upload id).
    try:
        rq1 = f"local-{int(upload_id)}"
        rq2 = f"local-smart-{int(upload_id)}"
        n = (
            db.query(Job)
            .filter(Job.organization_id == org, Job.rq_id.in_([rq1, rq2]))
            .delete(synchronize_session=False)
        )
        if n:
            deleted["jobs"] = int(n)
    except Exception:
        pass

    # Audit log: who deleted
    if get_settings().upload_audit_enabled:
        _safe_add_upload_audit_log(
            db,
            organization_id=org,
            upload_id=int(upload_id),
            actor_id=current_user.id,
            action="deleted",
            details=json.dumps({"deleted": deleted}),
        )

    # Finally delete upload record itself.
    try:
        db.delete(upload)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    return {"ok": True, "deleted": deleted, "upload_id": int(upload_id)}


_CATEGORY_COLOR_PALETTE = [
    "#ef4444",  # red
    "#f97316",  # orange
    "#eab308",  # amber
    "#22c55e",  # green
    "#06b6d4",  # cyan
    "#3b82f6",  # blue
    "#8b5cf6",  # violet
    "#ec4899",  # pink
]


def _map_category_to_activity_type(category: str) -> ActivityType:
    mapping = {
        "VERKAUFSFOERDERUNG": ActivityType.sales,
        "IMAGE": ActivityType.branding,
        "EMPLOYER_BRANDING": ActivityType.employer_branding,
        "KUNDENPFLEGE": ActivityType.kundenpflege,
    }
    return mapping.get((category or "").upper(), ActivityType.sales)


_ACTIVITY_CATEGORY_HEADER_ALIASES = (
    "category",
    "type",
    "kategorie",
    "bereich",
    "thema",
    "channel",
    "kanal",
    "kostenart",
    "cost_type",
    "costtype",
)


def _parse_csv_to_activities(content: bytes) -> List[Dict[str, Any]]:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for row in reader:
        rows.append(row)
    return rows


def _is_blankish(v: Any) -> bool:
    if v is None:
        return True
    try:
        s = str(v).strip().lower()
    except Exception:
        return False
    if s == "":
        return True
    return s in {"—", "–", "n/a", "na", "null", "none", "undefined"}


def _norm_header(s: str) -> str:
    s2 = str(s or "").strip().lower()
    # Normalize German chars for matching (ß->ss, umlauts->ae/oe/ue)
    s2 = (
        s2.replace("ß", "ss")
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
    )
    # Replace any non-alphanumeric with underscores, collapse repeats
    s2 = re.sub(r"[^a-z0-9]+", "_", s2)
    s2 = re.sub(r"_+", "_", s2).strip("_")
    return s2


def _suggest_header(headers_in: List[str], *names: str) -> Optional[str]:
    """
    Fuzzy header matcher:
    - normalizes punctuation (e.g. "Kosten (CHF)" -> "kosten_chf")
    - matches by exact, token, prefix or substring (handles plurals like "massnahmen")
    """
    prepared: List[tuple[str, str, List[str]]] = []
    for h in headers_in:
        hn = _norm_header(h)
        parts = [p for p in hn.split("_") if p]
        prepared.append((h, hn, parts))

    for n in names:
        n2 = _norm_header(n)
        if not n2:
            continue
        for orig, hn, parts in prepared:
            if not hn:
                continue
            if hn == n2:
                return orig
            if n2 in hn or hn in n2:
                return orig
            for p in parts:
                if p == n2 or p.startswith(n2) or n2.startswith(p):
                    return orig
    return None

def _dedupe_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for h in headers:
        base = str(h or "").strip() or "col"
        key = base
        n = seen.get(key, 0) + 1
        seen[key] = n
        out.append(key if n == 1 else f"{key}_{n}")
    return out

def _extract_xlsx_table_from_ws(ws: Any) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Best-effort extraction from "messy" Excel exports:
    - Detect a likely header row in the first N rows (titles / multi-row headers are common)
    - Normalize empty/duplicate headers
    - Skip repeated header rows and fully empty rows
    """
    # Scan first rows to find header row
    scan_rows = 40
    known_tokens = {
        # generic
        "title", "name", "beschreibung", "beschreibung/notes", "notes", "notiz", "kommentar",
        "massnahme", "maßnahme", "aktion", "initiative", "kampagne", "projekt",
        # dates
        "start", "beginn", "von", "ab", "startdatum", "start_date",
        "end", "ende", "bis", "enddatum", "end_date",
        # activity fields
        "category", "kategorie", "type", "bereich", "kanal", "format", "kostenart",
        "status", "phase", "workflow",
        "budget", "budgetchf", "kosten", "chf", "betrag",
        "owner", "verantwortlich", "zuständig",
        "priority", "prio", "gewicht", "weight",
        # crm-ish (so we still pick a good header)
        "company", "firma", "unternehmen", "email", "telefon", "website", "deal", "opportunity",
    }

    candidates: List[tuple[int, int, int]] = []  # (score, non_empty, row_idx)
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1):
        cells = [str(v or "").strip() for v in (row or [])]
        non_empty = sum(1 for c in cells if c)
        if non_empty < 3:
            continue
        hits = 0
        for c in cells:
            if not c:
                continue
            n = _norm_header(c).replace("__", "_")
            # split composite headers
            parts = [p for p in n.replace("/", "_").split("_") if p]
            if any(p in known_tokens for p in parts):
                hits += 1
        score = hits * 10 + non_empty
        candidates.append((score, non_empty, idx))

    header_row_idx = max(candidates, default=(0, 0, 1))[2]
    header_vals = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    raw_headers = [str(v or "").strip() for v in (header_vals or [])]
    # Fill empties and dedupe
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(raw_headers)]
    headers = _dedupe_headers(headers)

    rows_out: List[Dict[str, Any]] = []
    header_norm = [_norm_header(h) for h in headers]
    empty_streak = 0
    max_rows = 5000

    for r_idx, r in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=0):
        if r_idx >= max_rows:
            break
        vals = list(r or [])
        # pad to header length
        if len(vals) < len(headers):
            vals = vals + [None] * (len(headers) - len(vals))
        row_dict = {headers[i]: (vals[i] if i < len(vals) else None) for i in range(len(headers))}

        # Detect full-empty rows
        non_empty = 0
        for v in row_dict.values():
            if v is None:
                continue
            s = str(v).strip()
            if s:
                non_empty += 1
        if non_empty == 0:
            empty_streak += 1
            if empty_streak >= 25 and len(rows_out) > 0:
                break
            continue
        empty_streak = 0

        # Skip repeated header rows inside the sheet
        try:
            row_as_headers = [_norm_header(str(row_dict.get(headers[i]) or "")) for i in range(len(headers))]
            if sum(1 for i in range(len(headers)) if row_as_headers[i] and row_as_headers[i] == header_norm[i]) >= max(3, len(headers) // 3):
                continue
        except Exception:
            pass

        rows_out.append(row_dict)

    return headers, rows_out


def _extract_xlsx_table(content: bytes) -> tuple[List[str], List[Dict[str, Any]]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        raise HTTPException(status_code=415, detail="Excel (.xlsx) not supported without openpyxl. Please upload CSV.")
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    return _extract_xlsx_table_from_ws(wb.active)


def _extract_xlsx_tables(content: bytes, *, max_sheets: int = 8) -> List[Dict[str, Any]]:
    """
    Extract one best-effort table per worksheet.
    Returns: [{sheet, headers, rows}]
    """
    try:
        import openpyxl  # type: ignore
    except Exception:
        raise HTTPException(status_code=415, detail="Excel (.xlsx) not supported without openpyxl. Please upload CSV.")
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    def _try_extract_sap_mediaplan(ws: Any) -> Optional[Dict[str, Any]]:
        """
        SAP Magazin media plan format encodes the schedule as cell fill colors (blue blocks).
        We normalize it into a clean table with derived fields usable by the existing Activities importer.
        """
        try:
            # Find header row containing "Medium" and "Form"
            header_row: Optional[int] = None
            for r in range(1, min(int(getattr(ws, "max_row", 0) or 0), 80) + 1):
                found_medium = False
                found_form = False
                for c in range(1, min(int(getattr(ws, "max_column", 0) or 0), 40) + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    s = str(v).strip().lower()
                    if s == "medium":
                        found_medium = True
                    elif s == "form":
                        found_form = True
                if found_medium and found_form:
                    header_row = r
                    break

            if not header_row:
                return None

            def _find_col(label: str) -> Optional[int]:
                for c in range(1, int(getattr(ws, "max_column", 0) or 0) + 1):
                    v = ws.cell(header_row, c).value
                    if v is None:
                        continue
                    if str(v).strip().lower() == label.lower():
                        return c
                return None

            medium_col = _find_col("Medium")
            form_col = _find_col("Form")
            if not medium_col or not form_col:
                return None

            # Find CHF columns (two adjacent "CHF")
            chf_cols: List[int] = []
            for c in range(1, int(getattr(ws, "max_column", 0) or 0) + 1):
                v = ws.cell(header_row, c).value
                if v is None:
                    continue
                if str(v).strip().upper() == "CHF":
                    chf_cols.append(c)
            if len(chf_cols) < 1:
                return None

            timeline_start = max(chf_cols) + 1
            if timeline_start > int(getattr(ws, "max_column", 0) or 0):
                return None

            # Header rows above timeline
            year_row = max(1, header_row - 3)
            month_row = max(1, header_row - 2)
            day_row = header_row

            # month mapping (German + English + known typos)
            m_map = {
                "JANUAR": 1,
                "JANUARY": 1,
                "FEBRUAR": 2,
                "FEBRUARY": 2,
                "FEBURAR": 2,  # typo in the file
                "MÄRZ": 3,
                "MAERZ": 3,
                "MARCH": 3,
                "APRIL": 4,
                "MAI": 5,
                "MAY": 5,
                "JUNI": 6,
                "JUNE": 6,
                "JULI": 7,
                "JULY": 7,
                "AUGUST": 8,
                "SEPTEMBER": 9,
                "OKTOBER": 10,
                "OCTOBER": 10,
                "NOVEMBER": 11,
                "DEZEMBER": 12,
                "DECEMBER": 12,
            }

            def _norm_month(v: Any) -> str:
                s = str(v or "").strip().upper()
                # Handle umlauts and normalize to lookup keys
                s = s.replace("Ä", "AE").replace("Ö", "OE").replace("Ü", "UE")
                if s == "MAERZ":
                    return "MAERZ"
                # Map back common forms
                return s

            def _rgb(fill: Any) -> Optional[str]:
                try:
                    pat = getattr(fill, "patternType", None)
                    if not pat or pat == "none":
                        return None
                    fg = getattr(fill, "fgColor", None)
                    if not fg:
                        return None
                    if getattr(fg, "type", None) == "rgb":
                        rgb = getattr(fg, "rgb", None)
                        if isinstance(rgb, str) and len(rgb) == 8:
                            return rgb.upper()
                except Exception:
                    return None
                return None

            # Build date lookup per timeline column
            from datetime import date as _date

            col_date: Dict[int, _date] = {}
            last_year: Optional[int] = None
            last_month: Optional[int] = None
            for c in range(timeline_start, int(getattr(ws, "max_column", 0) or 0) + 1):
                yv = ws.cell(year_row, c).value
                mv = ws.cell(month_row, c).value
                dv = ws.cell(day_row, c).value

                # forward fill year/month across merged cells
                if yv not in (None, ""):
                    try:
                        last_year = int(str(yv).strip())
                    except Exception:
                        pass
                if mv not in (None, ""):
                    try:
                        # Convert "MAERZ" back to month 3; keep original umlaut values too
                        mkey = _norm_month(mv)
                        if mkey == "MAERZ":
                            last_month = 3
                        else:
                            # Try raw and normalized variants
                            raw = str(mv or "").strip().upper()
                            last_month = m_map.get(raw) or m_map.get(mkey)
                    except Exception:
                        pass

                try:
                    d = int(str(dv).strip()) if dv is not None and str(dv).strip().isdigit() else None
                except Exception:
                    d = None

                if last_year and last_month and d:
                    try:
                        col_date[c] = _date(last_year, last_month, d)
                    except Exception:
                        pass

            # Determine active fill color (prefer known SAP blue)
            from collections import Counter

            colors = Counter()
            for r in range(header_row + 1, int(getattr(ws, "max_row", 0) or 0) + 1):
                med = ws.cell(r, medium_col).value
                if med is None:
                    continue
                for c in range(timeline_start, int(getattr(ws, "max_column", 0) or 0) + 1):
                    rgb = _rgb(ws.cell(r, c).fill)
                    if rgb:
                        colors[rgb] += 1

            active_rgb: Optional[str] = None
            if colors:
                active_rgb = "FF01B0F0" if "FF01B0F0" in colors else colors.most_common(1)[0][0]
            if not active_rgb:
                return None

            def _segments_for_row(r: int) -> List[tuple[int, int]]:
                seg: List[tuple[int, int]] = []
                on = False
                start: Optional[int] = None
                for c in range(timeline_start, int(getattr(ws, "max_column", 0) or 0) + 1):
                    rgb = _rgb(ws.cell(r, c).fill)
                    is_on = bool(rgb and rgb == active_rgb)
                    if is_on and not on:
                        on = True
                        start = c
                    if on and not is_on:
                        seg.append((start or c, c - 1))
                        on = False
                        start = None
                if on and start is not None:
                    seg.append((start, int(getattr(ws, "max_column", 0) or start)))
                return seg

            # Metric columns
            impressions_col = _find_col("Impressions")
            clicks_col = _find_col("Clicks")
            views_col = _find_col("Views Adobe")
            ctr_col = _find_col("CTR")

            chf_2022_col = chf_cols[0]
            chf_2023_col = chf_cols[1] if len(chf_cols) > 1 else None

            headers_out = [
                "market",
                "section",
                "medium",
                "form",
                "impressions",
                "clicks",
                "views_adobe",
                "ctr",
                "chf_2022",
                "chf_2023",
                "budget_chf",
                "start",
                "end",
                "title",
                "notes",
            ]

            rows_out: List[Dict[str, Any]] = []
            current_market: Optional[str] = None
            current_section: Optional[str] = None

            for r in range(header_row + 1, int(getattr(ws, "max_row", 0) or 0) + 1):
                medium = ws.cell(r, medium_col).value
                form = ws.cell(r, form_col).value

                med_s = str(medium or "").strip()
                form_s = str(form or "").strip()
                if not med_s and not form_s:
                    continue

                # Section / market heading rows
                if med_s and not form_s:
                    low = med_s.lower()
                    if "deutsche schweiz" in low:
                        current_market = "Deutsche Schweiz"
                        current_section = med_s
                        continue
                    if "franz" in low and "schweiz" in low:
                        current_market = "Französische Schweiz"
                        current_section = med_s
                        continue
                    current_section = med_s
                    continue

                # Skip totals / summary rows
                low = med_s.lower()
                if low.startswith("total") or low.startswith("gesamttotal") or low.startswith("reserve"):
                    continue

                seg = _segments_for_row(r)
                start_dt = col_date.get(seg[0][0]) if seg else None
                end_dt = col_date.get(seg[-1][1]) if seg else None
                if start_dt and not end_dt:
                    end_dt = start_dt

                impressions = ws.cell(r, impressions_col).value if impressions_col else None
                clicks = ws.cell(r, clicks_col).value if clicks_col else None
                views = ws.cell(r, views_col).value if views_col else None
                ctr = ws.cell(r, ctr_col).value if ctr_col else None

                chf_2022 = ws.cell(r, chf_2022_col).value if chf_2022_col else None
                chf_2023 = ws.cell(r, chf_2023_col).value if chf_2023_col else None

                b1 = _parse_float(chf_2022) or 0.0
                b2 = _parse_float(chf_2023) or 0.0
                budget = b1 + b2

                title = f"{med_s} — {form_s}".strip(" —")
                notes = (
                    f"Market: {current_market or ''}\n"
                    f"Section: {current_section or ''}\n"
                    f"Medium: {med_s}\nForm: {form_s}\n"
                    f"Impressions: {impressions or ''}\nClicks: {clicks or ''}\nViews: {views or ''}\nCTR: {ctr or ''}\n"
                    f"Schedule: {(start_dt.isoformat() if start_dt else '')} → {(end_dt.isoformat() if end_dt else '')}"
                ).strip()

                rows_out.append(
                    {
                        "market": current_market,
                        "section": current_section,
                        "medium": med_s,
                        "form": form_s,
                        "impressions": impressions,
                        "clicks": clicks,
                        "views_adobe": views,
                        "ctr": ctr,
                        "chf_2022": chf_2022,
                        "chf_2023": chf_2023,
                        "budget_chf": budget if budget else None,
                        "start": start_dt.isoformat() if start_dt else None,
                        "end": end_dt.isoformat() if end_dt else None,
                        "title": title,
                        "notes": notes,
                    }
                )

            if not rows_out:
                return None

            return {"sheet": str(getattr(ws, "title", "") or "Sheet"), "headers": headers_out, "rows": rows_out}
        except Exception:
            return None

    out: List[Dict[str, Any]] = []
    for ws in wb.worksheets[:max_sheets]:
        special = _try_extract_sap_mediaplan(ws)
        if special:
            out.append(special)
            continue
        headers, rows = _extract_xlsx_table_from_ws(ws)
        if headers and rows:
            out.append({"sheet": str(getattr(ws, "title", "") or "Sheet"), "headers": headers, "rows": rows})
    if not out:
        # fall back to active even if empty, to show headers
        headers, rows = _extract_xlsx_table_from_ws(wb.active)
        out.append({"sheet": str(getattr(wb.active, "title", "") or "Sheet"), "headers": headers, "rows": rows})
    return out


def _suggest_mapping_crm(headers_in: List[str]) -> Dict[str, Optional[str]]:
    return {
        # Company
        "company_name": _suggest_header(headers_in, "company", "company_name", "firma", "unternehmen", "kundename", "name"),
        "company_website": _suggest_header(headers_in, "website", "web", "url", "domain"),
        "company_industry": _suggest_header(headers_in, "industry", "branche"),
        "company_email": _suggest_header(headers_in, "company_email", "email", "e-mail", "e_mail"),
        "company_phone": _suggest_header(headers_in, "phone", "telefon", "tel", "handy"),
        "company_notes": _suggest_header(headers_in, "notes", "notiz", "bemerkung", "kommentar"),
        # Contact
        "contact_name": _suggest_header(headers_in, "contact", "contact_name", "ansprechpartner", "kontakt", "kontakt_name"),
        "contact_email": _suggest_header(headers_in, "contact_email", "ansprechpartner_email", "kontakt_email"),
        "contact_phone": _suggest_header(headers_in, "contact_phone", "ansprechpartner_phone", "kontakt_phone", "telefon", "handy"),
        "contact_position": _suggest_header(headers_in, "position", "rolle", "funktion", "title"),
        # Deal
        "deal_title": _suggest_header(headers_in, "deal", "deal_title", "opportunity", "projekt", "project", "angebot"),
        "deal_value": _suggest_header(headers_in, "value", "betrag", "amount", "sum", "umsatz", "revenue", "chf", "preis"),
        "deal_stage": _suggest_header(headers_in, "stage", "status", "phase"),
        "deal_probability": _suggest_header(headers_in, "probability", "chance", "wahrscheinlichkeit"),
        "deal_expected_close_date": _suggest_header(headers_in, "expected_close_date", "close_date", "abschluss", "abschlussdatum"),
        "deal_owner": _suggest_header(headers_in, "owner", "deal_owner", "verantwortlich", "owner_email"),
        "deal_notes": _suggest_header(headers_in, "deal_notes", "notes_deal", "bemerkung_deal", "kommentar"),
    }

def _suggest_mapping_content(headers_in: List[str]) -> Dict[str, Optional[str]]:
    return {
        "title": _suggest_header(headers_in, "title", "name", "titel", "thema", "betreff", "subject", "aufgabe", "todo", "task", "massnahme", "massnahmen", "aktion"),
        "channel": _suggest_header(headers_in, "channel", "kanal", "plattform", "platform"),
        "format": _suggest_header(headers_in, "format", "type", "typ", "content_type"),
        "status": _suggest_header(headers_in, "status", "phase", "workflow"),
        "due_at": _suggest_header(headers_in, "due", "due_at", "deadline", "faellig", "fällig", "abgabe", "abgabedatum", "due_date"),
        "scheduled_at": _suggest_header(
            headers_in,
            "scheduled",
            "scheduled_at",
            "publish",
            "publishing",
            "posting",
            "post_date",
            "publishing_date",
            "termin",
            "datum",
        ),
        "tags": _suggest_header(headers_in, "tags", "tag", "labels", "label"),
        "brief": _suggest_header(headers_in, "brief", "beschreibung", "notes", "notiz", "kommentar", "bemerkung"),
        "body": _suggest_header(headers_in, "body", "text", "copy", "inhalt", "content", "post"),
        "language": _suggest_header(headers_in, "language", "sprache", "lang"),
        "tone": _suggest_header(headers_in, "tone", "stil", "tonalitaet", "tonalität"),
        "owner_email": _suggest_header(headers_in, "owner", "owner_email", "verantwortlich", "zustaendig", "zuständig", "assigned_to", "assignee"),
    }


def _suggest_mapping_budget(headers_in: List[str]) -> Dict[str, Optional[str]]:
    return {
        "period": _suggest_header(headers_in, "period", "quartal", "quarter", "q", "jahr_quartal", "year_quarter", "jahr", "year"),
        "category": _suggest_header(headers_in, "category", "kategorie", "bereich", "thema", "type", "kpi_kategorie"),
        "amount": _suggest_header(headers_in, "amount", "budget", "budget_chf", "budgetchf", "kosten", "betrag", "chf", "summe", "spend", "preis", "kosten_chf"),
        "metric": _suggest_header(headers_in, "metric", "kpi", "kennzahl", "ziel_kpi", "kpi_name", "name"),
        "target": _suggest_header(headers_in, "target", "zielwert", "target_value", "wert", "value", "ziel"),
        "unit": _suggest_header(headers_in, "unit", "einheit", "currency"),
    }


def _suggest_mapping_activities(headers_in: List[str]) -> Dict[str, Optional[str]]:
    return {
        "title": _suggest_header(headers_in, "title", "name", "massnahme", "massnahmen", "maßnahme", "maßnahmen", "aufgabe", "todo", "task", "aktion", "initiative", "kampagne", "projekt"),
        "category": _suggest_header(headers_in, *_ACTIVITY_CATEGORY_HEADER_ALIASES),
        "status": _suggest_header(headers_in, "status", "phase", "workflow"),
        "budget": _suggest_header(headers_in, "budget", "budgetchf", "budget_chf", "kosten", "betrag", "chf", "summe", "preis"),
        "notes": _suggest_header(headers_in, "notes", "expected_output", "beschreibung", "kommentar", "notiz", "bemerkung"),
        "start": _suggest_header(headers_in, "start", "start_date", "beginn", "von", "ab", "startdatum", "datum_start"),
        "end": _suggest_header(headers_in, "end", "end_date", "ende", "bis", "enddatum", "datum_ende"),
        "weight": _suggest_header(headers_in, "weight", "gewicht", "prio", "priority"),
    }


def _parse_float(v: Any) -> Optional[float]:
    if _is_blankish(v):
        return None
    try:
        s = str(v).strip().replace("'", "").replace(" ", "")
        # Accept 1'234.50 or 1.234,50
        if "," in s and "." in s:
            # choose last separator as decimal
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s and "." not in s:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return None


def _parse_percent(v: Any) -> Optional[float]:
    """
    Parse percent-ish values and return a percentage number (0..100+).
    Accepts: 0.12, 12, '12%', '12,5 %', "0,12", etc.
    """
    if _is_blankish(v):
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        s = s.replace("%", "").replace("'", "").replace(" ", "")
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        val = float(s)
        # Heuristic: 0..1 likely means fraction
        if 0 <= val <= 1:
            val = val * 100.0
        return val
    except Exception:
        # Fallback: reuse float parser (will handle 1'234.50 patterns)
        try:
            f = _parse_float(str(v).replace("%", ""))
            if f is None:
                return None
            return f * 100.0 if 0 <= f <= 1 else f
        except Exception:
            return None


def _parse_int(v: Any) -> Optional[int]:
    if _is_blankish(v):
        return None
    try:
        return int(float(str(v).strip().replace("%", "")))
    except Exception:
        return None


def _parse_datetime_loose(v: Any):
    if _is_blankish(v):
        return None
    try:
        from datetime import datetime, date

        if isinstance(v, datetime):
            return v
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day)
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None
    except Exception:
        return None


def _current_period() -> str:
    from datetime import datetime

    now = datetime.utcnow()
    q = (now.month - 1) // 3 + 1
    return f"{now.year}-Q{q}"


def _norm_budget_category(v: Any) -> str:
    s = str(v or "").strip()
    if not s:
        return "VERKAUFSFOERDERUNG"
    up = s.upper().replace(" ", "_").replace("-", "_")
    alias = {
        "SALES": "VERKAUFSFOERDERUNG",
        "SALES_PROMO": "VERKAUFSFOERDERUNG",
        "VERKAUF": "VERKAUFSFOERDERUNG",
        "PROMOTION": "VERKAUFSFOERDERUNG",
        "BRAND": "IMAGE",
        "BRANDING": "IMAGE",
        "EMPLOYER": "EMPLOYER_BRANDING",
        "EMPLOYERBRANDING": "EMPLOYER_BRANDING",
        "HR": "EMPLOYER_BRANDING",
        "RETENTION": "KUNDENPFLEGE",
        "CUSTOMER": "KUNDENPFLEGE",
        "PFLEGE": "KUNDENPFLEGE",
    }
    return alias.get(up, up)


def _parse_tags(v: Any) -> Optional[List[str]]:
    if _is_blankish(v):
        return None
    s = str(v).strip()
    if not s:
        return None
    parts = [p.strip() for p in s.replace("|", ",").replace(";", ",").split(",")]
    out: List[str] = []
    seen = set()
    for p in parts:
        if not p:
            continue
        key = p.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out or None


@router.get("")
def list_uploads(
    db: Session = Depends(get_db_session),
    current_user: User = Depends(get_current_user),
):
    # Uploads are global in current schema; do not show them in demo mode to avoid leaking real data.
    if is_demo_user(current_user):
        return {"items": []}
    org = get_org_id(current_user)
    can_manage_all = current_user.role in {UserRole.admin, UserRole.editor}
    q = db.query(Upload).filter(Upload.organization_id == org)
    if not can_manage_all:
        # Regular users can only see their own uploads (avoid cross-user leakage inside org).
        q = q.filter(Upload.owner_id == current_user.id)
    items = q.order_by(Upload.created_at.desc()).all()
    return {
        "items": [
            {
                "id": str(u.id),
                "original_name": u.original_name,
                "file_type": u.file_type,
                "file_size": int(u.file_size or 0),
                "stored_in_db": bool(getattr(u, "stored_in_db", False)),
                "sha256": getattr(u, "sha256", None),
                "created_at": u.created_at,
            }
            for u in items
        ]
    }


@router.post("")
def upload_file(
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(default=None),
    import_kind: Optional[str] = Form(default="activities"),
    db: Session = Depends(get_db_session),
    current_user: User = Depends(require_writable_user),
):
    """
    Accept a file upload.

    - CSV/XLSX: import rows as Activities (with optional mapping)
    - Other file types: store file in DB (no import)

    Columns supported (case-insensitive):
    title, category|type, status, weight, budget|budgetCHF, notes,
    start|start_date, end|end_date
    """
    settings = get_settings()
    org = get_org_id(current_user)

    # Save upload metadata (and optionally bytes in DB)
    upload = Upload(
        original_name=file.filename or "file",
        file_type=file.content_type or "",
        file_size=0,
        organization_id=org,
        owner_id=current_user.id,
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)

    # Read all content
    content = file.file.read() or b""
    if len(content) > settings.upload_max_bytes:
        # Avoid silently truncating / losing data.
        raise HTTPException(
            status_code=413,
            detail=f"File too large for current plan (max {settings.upload_max_bytes} bytes).",
        )
    upload.file_size = len(content or b"")
    upload.sha256 = hashlib.sha256(content).hexdigest()
    upload.stored_in_db = bool(settings.upload_store_in_db)
    if settings.upload_store_in_db:
        upload.content = content
    db.add(upload)
    db.commit()
    db.refresh(upload)

    created_count = 0
    skipped_count = 0

    # Parse CSV or Excel
    rows: List[Dict[str, Any]] = []
    ctype = (file.content_type or "").lower()
    name_lower = (file.filename or "").lower()

    try:
        if ctype in {"text/csv", "application/csv", "text/plain"} or name_lower.endswith(".csv"):
            rows = _parse_csv_to_activities(content)
        elif name_lower.endswith(".xlsx") or ctype in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
            # Use multi-sheet extraction with format-specific normalizers (e.g., SAP Mediaplan timeline fills).
            tables = _extract_xlsx_tables(content)
            picked = None
            for t in tables:
                if str(t.get("sheet") or "").strip().lower() == "mediaplan":
                    picked = t
                    break
            if not picked and tables:
                picked = max(tables, key=lambda t: len(t.get("rows") or []))
            rows = list((picked or {}).get("rows") or [])
        else:
            # Non-tabular file: keep as stored upload only (no import)
            return {
                "ok": True,
                "item": {
                    "id": str(upload.id),
                    "original_name": upload.original_name,
                    "file_type": upload.file_type,
                    "file_size": int(upload.file_size or 0),
                    "created_at": upload.created_at,
                },
                "import": {"created": 0, "skipped": 0, "mode": "stored"},
            }

        # Normalize header names
        def g(row: Dict[str, Any], *keys: str):
            for k in keys:
                for variant in (k, k.lower(), k.upper(), k.replace(" ", "_")):
                    if variant in row and not _is_blankish(row[variant]):
                        return row[variant]
            return None

        from datetime import datetime, date

        def parse_date(v):
            if _is_blankish(v): 
                return None
            if isinstance(v, (datetime, date)): return v if isinstance(v, date) else v.date()
            s = str(v).strip()
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            try:
                # ISO
                return datetime.fromisoformat(s.replace('Z', '+00:00')).date()
            except Exception:
                return None

        # Optional mapping from form JSON { field -> headerName }
        mapping_dict: Dict[str, Any] = {}
        if mapping:
            try:
                import json as _json
                mapping_dict = _json.loads(mapping)
            except Exception:
                mapping_dict = {}

        # Optional: category value remap { "DIGITAL_MARKETING": "VERKAUFSFOERDERUNG", ... }
        raw_category_value_map: Dict[str, Any] = {}
        try:
            cvm = (
                mapping_dict.get("category_value_map")
                or mapping_dict.get("__category_value_map")
                or mapping_dict.get("categoryValueMap")
            )
            if isinstance(cvm, dict):
                raw_category_value_map = cvm
        except Exception:
            raw_category_value_map = {}

        def remap_category_value(v: Any) -> str:
            s = str(v or "").strip()
            if not s:
                return s
            key = s.upper()
            # Compare case-insensitively; frontend sends normalized keys by default
            for k, target in raw_category_value_map.items():
                try:
                    if str(k).strip().upper() == key:
                        return str(target)
                except Exception:
                    continue
            return s

        def gm(row: Dict[str, Any], field: str, *fallbacks: str):
            header = (mapping_dict.get(field) or '').strip()
            if header:
                v = row.get(header)
                if not _is_blankish(v):
                    return v
            return g(row, *fallbacks)

        kind = (import_kind or "activities").strip().lower()
        if kind not in {"activities", "crm", "content", "budget"}:
            kind = "activities"

        if kind == "crm":
            # Restrict CRM import to editor/admin to prevent accidental org-wide writes.
            if current_user.role not in {UserRole.admin, UserRole.editor}:
                raise HTTPException(status_code=403, detail="CRM import requires editor/admin")

            for row in rows:
                try:
                    company_name = str(gm(row, "company_name", "company", "firma", "unternehmen", "name") or "").strip()
                    if not company_name:
                        skipped_count += 1
                        continue

                    company = (
                        db.query(Company)
                        .filter(Company.organization_id == org, Company.name == company_name)
                        .first()
                    )
                    if not company:
                        company = Company(
                            organization_id=org,
                            name=company_name,
                            source_upload_id=int(upload.id),
                        )
                        db.add(company)
                        db.flush()
                        created_count += 1
                    # Best-effort enrichment (only set when provided)
                    website = gm(row, "company_website", "website", "domain", "url")
                    if website not in (None, ""):
                        company.website = str(website).strip()
                    industry = gm(row, "company_industry", "industry", "branche")
                    if industry not in (None, ""):
                        company.industry = str(industry).strip()
                    c_email = gm(row, "company_email", "email")
                    if c_email not in (None, ""):
                        company.email = str(c_email).strip()
                    c_phone = gm(row, "company_phone", "phone", "telefon", "tel")
                    if c_phone not in (None, ""):
                        company.phone = str(c_phone).strip()
                    c_notes = gm(row, "company_notes", "notes", "notiz", "bemerkung")
                    if c_notes not in (None, ""):
                        company.notes = str(c_notes).strip()[:1024]
                    db.add(company)

                    # Contact (optional)
                    contact_email = gm(row, "contact_email", "kontakt_email", "ansprechpartner_email")
                    contact_name = gm(row, "contact_name", "kontakt", "ansprechpartner", "contact")
                    contact = None
                    if contact_email not in (None, "") or contact_name not in (None, ""):
                        ce = str(contact_email or "").strip().lower()
                        cn = str(contact_name or "").strip()
                        q = db.query(Contact).filter(Contact.organization_id == org, Contact.company_id == company.id)
                        if ce:
                            contact = q.filter(Contact.email == ce).first()
                        if not contact and cn:
                            contact = q.filter(Contact.name == cn).first()
                        if not contact:
                            contact = Contact(organization_id=org, company_id=company.id, name=cn or ce or "Contact")
                            if ce:
                                contact.email = ce
                            contact.source_upload_id = int(upload.id)
                            db.add(contact)
                            db.flush()
                            created_count += 1
                        # Enrich
                        if cn:
                            contact.name = cn
                        if ce:
                            contact.email = ce
                        phone = gm(row, "contact_phone", "telefon", "phone")
                        if phone not in (None, ""):
                            contact.phone = str(phone).strip()
                        pos = gm(row, "contact_position", "position", "rolle", "funktion")
                        if pos not in (None, ""):
                            contact.position = str(pos).strip()
                        db.add(contact)

                    # Deal (optional)
                    deal_title = gm(row, "deal_title", "deal", "opportunity", "projekt", "project", "angebot")
                    if deal_title not in (None, ""):
                        dt = str(deal_title).strip()
                        if dt:
                            deal = (
                                db.query(Deal)
                                .filter(Deal.organization_id == org, Deal.company_id == company.id, Deal.title == dt)
                                .first()
                            )
                            if not deal:
                                deal = Deal(
                                    organization_id=org,
                                    company_id=company.id,
                                    title=dt,
                                    owner=str(gm(row, "deal_owner", "owner", "verantwortlich") or (current_user.email or "owner")),
                                    source_upload_id=int(upload.id),
                                )
                                db.add(deal)
                                db.flush()
                                created_count += 1
                            value = _parse_float(gm(row, "deal_value", "value", "betrag", "amount", "chf"))
                            if value is not None:
                                deal.value = value
                            stage = gm(row, "deal_stage", "stage", "phase", "status")
                            if stage not in (None, ""):
                                deal.stage = str(stage).strip().lower()[:30]
                            prob = _parse_int(gm(row, "deal_probability", "probability", "chance", "wahrscheinlichkeit"))
                            if prob is not None:
                                deal.probability = max(0, min(100, prob))
                            close_dt = _parse_datetime_loose(gm(row, "deal_expected_close_date", "expected_close_date", "close_date", "abschlussdatum"))
                            if close_dt:
                                deal.expected_close_date = close_dt
                            dnotes = gm(row, "deal_notes", "notes")
                            if dnotes not in (None, ""):
                                deal.notes = str(dnotes).strip()[:1024]
                            # Attach optional contact_id if present
                            if contact is not None:
                                deal.contact_id = contact.id
                            db.add(deal)
                except Exception:
                    skipped_count += 1

            db.commit()
        elif kind == "budget":
            if current_user.role not in {UserRole.admin, UserRole.editor}:
                raise HTTPException(status_code=403, detail="Budget import requires editor/admin")

            default_period = _current_period()
            for row in rows:
                try:
                    period = str(gm(row, "period", "period", "quartal", "quarter", "jahr_quartal", "year_quarter") or default_period).strip()
                    if not period:
                        period = default_period

                    # Budget target row
                    cat_raw = gm(row, "category", "category", "kategorie", "bereich", "thema", "type")
                    amt_raw = gm(row, "amount", "amount", "budget", "budgetchf", "budget_chf", "kosten", "betrag", "chf", "summe")
                    amount = _parse_float(amt_raw)
                    if cat_raw not in (None, "") and amount is not None:
                        category = _norm_budget_category(cat_raw)
                        existing = (
                            db.query(BudgetTarget)
                            .filter(
                                BudgetTarget.organization_id == org,
                                BudgetTarget.period == period,
                                BudgetTarget.category == category,
                            )
                            .first()
                        )
                        if existing:
                            existing.amount_chf = float(amount)
                            existing.source_upload_id = int(upload.id)
                            db.add(existing)
                        else:
                            db.add(BudgetTarget(organization_id=org, period=period, category=category, amount_chf=float(amount), source_upload_id=int(upload.id)))
                            created_count += 1

                    # KPI target row
                    metric = gm(row, "metric", "metric", "kpi", "kennzahl", "ziel_kpi", "kpi_name", "name")
                    target_val = _parse_float(gm(row, "target", "target", "zielwert", "target_value", "wert", "value"))
                    unit = gm(row, "unit", "unit", "einheit", "currency")
                    if metric not in (None, "") and target_val is not None:
                        m = str(metric).strip()
                        if m:
                            existing_k = (
                                db.query(KpiTarget)
                                .filter(KpiTarget.organization_id == org, KpiTarget.period == period, KpiTarget.metric == m)
                                .first()
                            )
                            if existing_k:
                                existing_k.target_value = float(target_val)
                                existing_k.unit = str(unit).strip()[:20] if unit not in (None, "") else existing_k.unit
                                existing_k.source_upload_id = int(upload.id)
                                db.add(existing_k)
                            else:
                                db.add(
                                    KpiTarget(
                                        organization_id=org,
                                        period=period,
                                        metric=m,
                                        target_value=float(target_val),
                                        unit=(str(unit).strip()[:20] if unit not in (None, "") else None),
                                        source_upload_id=int(upload.id),
                                    )
                                )
                                created_count += 1
                except Exception:
                    skipped_count += 1

            db.commit()
        elif kind == "content":
            # Any writable user can import content items into their organization.
            can_manage_all = current_user.role in {UserRole.admin, UserRole.editor}
            for row in rows:
                try:
                    title = str(gm(row, "title", "title", "name", "titel", "thema", "betreff") or "").strip()
                    if not title:
                        skipped_count += 1
                        continue

                    channel = str(gm(row, "channel", "channel", "kanal", "plattform", "platform") or "Website").strip()[:100]
                    fmt = gm(row, "format", "format", "type", "typ", "content_type")
                    fmt_s = str(fmt).strip()[:100] if fmt not in (None, "") else None

                    raw_status = str(gm(row, "status", "status", "phase", "workflow") or "").strip().lower()
                    status_map = {
                        "idea": ContentItemStatus.IDEA,
                        "idee": ContentItemStatus.IDEA,
                        "draft": ContentItemStatus.DRAFT,
                        "entwurf": ContentItemStatus.DRAFT,
                        "review": ContentItemStatus.REVIEW,
                        "prüfung": ContentItemStatus.REVIEW,
                        "pruefung": ContentItemStatus.REVIEW,
                        "approved": ContentItemStatus.APPROVED,
                        "freigegeben": ContentItemStatus.APPROVED,
                        "scheduled": ContentItemStatus.SCHEDULED,
                        "geplant": ContentItemStatus.SCHEDULED,
                        "published": ContentItemStatus.PUBLISHED,
                        "veröffentlicht": ContentItemStatus.PUBLISHED,
                        "veroeffentlicht": ContentItemStatus.PUBLISHED,
                        "archived": ContentItemStatus.ARCHIVED,
                        "archiviert": ContentItemStatus.ARCHIVED,
                        "blocked": ContentItemStatus.BLOCKED,
                        "blockiert": ContentItemStatus.BLOCKED,
                    }
                    status = status_map.get(raw_status, ContentItemStatus.DRAFT)

                    due_at = _parse_datetime_loose(gm(row, "due_at", "due_at", "due", "deadline", "fällig", "faellig", "abgabe", "due_date"))
                    scheduled_at = _parse_datetime_loose(
                        gm(
                            row,
                            "scheduled_at",
                            "scheduled_at",
                            "scheduled",
                            "publish",
                            "publishing_date",
                            "post_date",
                            "termin",
                            "datum",
                        )
                    )
                    tags = _parse_tags(gm(row, "tags", "tags", "labels", "label", "tag"))
                    brief = gm(row, "brief", "brief", "beschreibung", "notes", "notiz", "kommentar", "bemerkung")
                    body = gm(row, "body", "body", "text", "copy", "inhalt", "content", "post")
                    language = str(gm(row, "language", "language", "sprache", "lang") or "de").strip()[:10]
                    tone = gm(row, "tone", "tone", "stil", "tonalität", "tonalitaet")
                    tone_s = str(tone).strip()[:50] if tone not in (None, "") else None

                    desired_owner_id = current_user.id
                    owner_email = gm(row, "owner_email", "owner_email", "owner", "verantwortlich", "zuständig", "zustaendig", "assignee")
                    if can_manage_all and owner_email not in (None, ""):
                        oe = str(owner_email).strip().lower()
                        if oe:
                            u = db.query(User).filter(User.organization_id == org, User.email == oe).first()
                            if u:
                                desired_owner_id = u.id

                    item = ContentItem(
                        organization_id=org,
                        owner_id=desired_owner_id,
                        source_upload_id=int(upload.id),
                        title=title[:255],
                        channel=channel or "Website",
                        format=fmt_s,
                        status=status,
                        tags=tags,
                        brief=(str(brief).strip() if brief not in (None, "") else None),
                        body=(str(body).strip() if body not in (None, "") else None),
                        language=language or "de",
                        tone=tone_s,
                        due_at=due_at,
                        scheduled_at=scheduled_at,
                    )
                    db.add(item)
                    created_count += 1
                except Exception:
                    skipped_count += 1

            db.commit()
        else:
            for row in rows:
                try:
                    title = (
                        gm(row, "title", "title", "name", "massnahme", "maßnahme", "aktion", "initiative", "kampagne", "projekt")
                        or "Untitled"
                    )
                    category = gm(row, "category", *_ACTIVITY_CATEGORY_HEADER_ALIASES) or "VERKAUFSFOERDERUNG"
                    category = remap_category_value(category) or "VERKAUFSFOERDERUNG"
                    raw_status = (gm(row, "status", "status", "phase", "workflow") or "ACTIVE")
                    s_norm = str(raw_status or "").strip().lower()
                    status_map = {
                        "geplant": "PLANNED",
                        "planung": "PLANNED",
                        "planned": "PLANNED",
                        "aktiv": "ACTIVE",
                        "active": "ACTIVE",
                        "in_progress": "ACTIVE",
                        "in progress": "ACTIVE",
                        "laufend": "ACTIVE",
                        "done": "DONE",
                        "fertig": "DONE",
                        "abgeschlossen": "DONE",
                        "completed": "DONE",
                        "pausiert": "PAUSED",
                        "paused": "PAUSED",
                        "abgebrochen": "CANCELLED",
                        "cancelled": "CANCELLED",
                        "canceled": "CANCELLED",
                    }
                    status = status_map.get(s_norm, str(raw_status).upper() if raw_status not in (None, "") else "ACTIVE")

                    budget = _parse_float(gm(row, "budget", "budget", "budgetCHF", "kosten", "betrag", "chf"))
                    weight = _parse_float(gm(row, "weight", "weight", "gewicht", "prio", "priority"))
                    notes = gm(row, "notes", "notes", "expected_output", "beschreibung", "kommentar", "notiz", "bemerkung")
                    start = parse_date(gm(row, "start", "start", "start_date", "beginn", "von", "ab", "startdatum"))
                    end = parse_date(gm(row, "end", "end", "end_date", "ende", "bis", "enddatum"))

                    activity = Activity(
                        title=str(title),
                        type=_map_category_to_activity_type(str(category)),
                        category_name=str(category),
                        status=str(status).upper(),
                        source_upload_id=int(upload.id),
                        budget=budget,
                        weight=weight,
                        expected_output=str(notes) if notes not in (None, "") else None,
                        start_date=start,
                        end_date=end,
                        owner_id=current_user.id,
                        organization_id=org,
                    )
                    db.add(activity)
                    created_count += 1
                except Exception:
                    skipped_count += 1

        db.commit()

        # Record a completed job
        rq_id = f"local-{upload.id}"
        job = Job(
            rq_id=rq_id,
            type=("import_crm" if kind == "crm" else "import_activities"),
            status="finished",
            result=f"created={created_count};skipped={skipped_count}",
            organization_id=org,
        )
        db.add(job)
        db.commit()

        db.refresh(upload)
        return {
            "ok": True,
            "item": {
                "id": str(upload.id),
                "original_name": upload.original_name,
                "file_type": upload.file_type,
                "file_size": int(upload.file_size or 0),
                "created_at": upload.created_at,
            },
            "import": {"created": created_count, "skipped": skipped_count},
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/preview")
def preview_upload(
    file: UploadFile = File(...),
    import_kind: Optional[str] = Form(default="activities"),
    current_user: User = Depends(get_current_user),
):
    """Return headers, first rows and suggested mapping for CSV/XLSX."""
    content = file.file.read()
    ctype = (file.content_type or "").lower()
    name_lower = (file.filename or "").lower()

    rows: List[Dict[str, Any]] = []
    headers: List[str] = []
    suggested: Dict[str, Optional[str]] = {}
    category_values: set[str] = set()
    category_values_by_header: Dict[str, set[str]] = {}
    max_sample_rows = 5
    max_scan_rows = 5000
    max_unique_categories = 200
    normalized_activity_category_aliases = {_norm_header(v) for v in _ACTIVITY_CATEGORY_HEADER_ALIASES}

    def suggest_mapping(headers_in: List[str]) -> Dict[str, Optional[str]]:
        return _suggest_mapping_activities(headers_in)

    kind = (import_kind or "activities").strip().lower()
    if kind not in {"activities", "crm", "content", "budget"}:
        kind = "activities"

    if ctype in {"text/csv", "application/csv", "text/plain"} or name_lower.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        if kind == "crm":
            suggested = _suggest_mapping_crm(headers)
        elif kind == "content":
            suggested = _suggest_mapping_content(headers)
        elif kind == "budget":
            suggested = _suggest_mapping_budget(headers)
        else:
            suggested = suggest_mapping(headers)
        candidate_category_headers: List[str] = []
        if kind == "activities":
            for h in headers:
                if _norm_header(h) in normalized_activity_category_aliases:
                    candidate_category_headers.append(h)
            if suggested.get("category") and suggested.get("category") not in candidate_category_headers:
                candidate_category_headers.append(str(suggested.get("category")))
            for h in candidate_category_headers:
                category_values_by_header[h] = set()
        for i, row in enumerate(reader):
            if i < max_sample_rows:
                rows.append(row)
            for h in candidate_category_headers:
                v = row.get(h)
                if v in (None, ""):
                    continue
                s = str(v).strip()
                if not s:
                    continue
                category_values_by_header.setdefault(h, set()).add(s)
                if h == suggested.get("category"):
                    category_values.add(s)
            if i >= max_scan_rows:
                break
    elif name_lower.endswith(".xlsx") or ctype in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
        # Use multi-sheet extraction with format-specific normalizers (e.g., SAP Mediaplan timeline fills).
        tables = _extract_xlsx_tables(content)
        # Prefer a "Mediaplan" sheet if present; otherwise use the largest table.
        picked = None
        for t in tables:
            if str(t.get("sheet") or "").strip().lower() == "mediaplan":
                picked = t
                break
        if not picked and tables:
            picked = max(tables, key=lambda t: len(t.get("rows") or []))
        headers = list((picked or {}).get("headers") or [])
        parsed_rows = list((picked or {}).get("rows") or [])
        if kind == "crm":
            suggested = _suggest_mapping_crm(headers)
        elif kind == "content":
            suggested = _suggest_mapping_content(headers)
        elif kind == "budget":
            suggested = _suggest_mapping_budget(headers)
        else:
            suggested = suggest_mapping(headers)
        candidate_category_headers = []
        if kind == "activities":
            for h in headers:
                if _norm_header(h) in normalized_activity_category_aliases:
                    candidate_category_headers.append(h)
            if suggested.get("category") and suggested.get("category") not in candidate_category_headers:
                candidate_category_headers.append(str(suggested.get("category")))
            for h in candidate_category_headers:
                category_values_by_header[h] = set()
        for i, row in enumerate(parsed_rows):
            if i < max_sample_rows:
                rows.append(row)
            for h in candidate_category_headers:
                v = row.get(h)
                if v in (None, ""):
                    continue
                s = str(v).strip()
                if not s:
                    continue
                category_values_by_header.setdefault(h, set()).add(s)
                if h == suggested.get("category"):
                    category_values.add(s)
            if i >= max_scan_rows:
                break
    else:
        raise HTTPException(status_code=415, detail="Unsupported file type. Upload CSV or XLSX")

    return {
        "headers": headers,
        "samples": rows,
        "suggested_mapping": suggested,
        **({"category_values": sorted(list(category_values))} if kind == "activities" else {}),
        **(
            {
                "category_values_by_header": {
                    h: sorted(list(vals))[:max_unique_categories]
                    for h, vals in category_values_by_header.items()
                    if vals
                }
            }
            if kind == "activities"
            else {}
        ),
        "import_kind": kind,
    }


def _mask_value(v: Any) -> Any:
    """
    Light PII masking for AI analysis payloads.
    Keeps structure while avoiding leaking real emails/phones.
    """
    if v is None:
        return None
    try:
        s = str(v).strip()
    except Exception:
        return None
    if not s:
        return ""
    s_low = s.lower()
    if "@" in s and "." in s.split("@")[-1]:
        return "<email>"
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 8 and any(ch in s for ch in ["+", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]):
        # likely phone or id-like number
        return "<number>"
    if s_low.startswith("http://") or s_low.startswith("https://"):
        # keep only hostname
        try:
            from urllib.parse import urlparse

            host = urlparse(s).netloc
            return f"https://{host}" if host else "<url>"
        except Exception:
            return "<url>"
    # keep short strings, truncate long free-text
    if len(s) > 240:
        return s[:240] + "…"
    return s


def _safe_json_from_model_reply(reply: str) -> Dict[str, Any]:
    """
    OpenAI may return JSON wrapped in code fences or with surrounding text.
    Extract a JSON object robustly.
    """
    s = (reply or "").strip()
    if not s:
        raise ValueError("empty_reply")

    # Strip common ```json ... ``` wrappers
    if s.startswith("```"):
        lines = s.split("\n")
        if len(lines) >= 2:
            lines = lines[1:]
        s = "\n".join(lines)
        if s.endswith("```"):
            s = s[:-3]
        s = s.strip()

    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return obj
        raise ValueError("reply_not_object")
    except Exception:
        i = s.find("{")
        j = s.rfind("}")
        if i >= 0 and j > i:
            cand = s[i : j + 1]
            obj2 = json.loads(cand)
            if isinstance(obj2, dict):
                return obj2
        raise


def _compute_missingness(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    if not headers:
        return out
    n = max(1, len(rows))
    for h in headers[:80]:
        miss = 0
        for r in rows:
            if _is_blankish(r.get(h)):
                miss += 1
        out[h] = miss / n
    return out


def _top_values(rows: List[Dict[str, Any]], header: str, limit: int = 8) -> List[Dict[str, Any]]:
    from collections import Counter

    c: Counter[str] = Counter()
    for r in rows:
        v = r.get(header)
        if _is_blankish(v):
            continue
        s = str(v).strip()
        if not s:
            continue
        c[s] += 1
    return [{"value": k, "count": int(v)} for k, v in c.most_common(limit)]

def _column_stats(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Compute full-table statistics per column:
    - missing ratio
    - inferred type (number/date/text)
    - numeric min/max/avg
    - date min/max
    - top values (categorical)
    """
    from collections import Counter

    out: Dict[str, Any] = {"columns": {}, "row_count": len(rows)}
    if not headers or not rows:
        return out

    for h in headers[:120]:
        miss = 0
        num_ok = 0
        num_bad = 0
        date_ok = 0
        date_bad = 0
        nums: List[float] = []
        dates = []
        top = Counter()
        unique = set()
        unique_cap = 800

        for r in rows:
            v = r.get(h)
            if _is_blankish(v):
                miss += 1
                continue

            # track uniques (capped)
            try:
                s = str(v).strip()
            except Exception:
                s = ""
            if s and len(unique) < unique_cap:
                unique.add(s)
            if s and len(s) <= 80:
                top[s] += 1

            f = _parse_float(v)
            if f is not None:
                num_ok += 1
                nums.append(float(f))
            else:
                # only count as numeric bad if it looks number-ish
                if any(ch.isdigit() for ch in s):
                    num_bad += 1

            dt = _parse_datetime_loose(v)
            if dt is not None:
                date_ok += 1
                dates.append(dt)
            else:
                # count as date bad if it contains common separators
                if any(ch in s for ch in [".", "-", "/"]) and any(ch.isdigit() for ch in s):
                    date_bad += 1

        n = max(1, len(rows))
        missing_ratio = miss / n

        # Infer type by dominance
        typ = "text"
        if num_ok >= max(3, int(0.5 * (n - miss))):
            typ = "number"
        elif date_ok >= max(3, int(0.4 * (n - miss))):
            typ = "date"

        col: Dict[str, Any] = {
            "missing_ratio": missing_ratio,
            "inferred_type": typ,
            "unique_sampled": len(unique),
            "top_values": [{"value": k, "count": int(v)} for k, v in top.most_common(10)],
        }
        if nums:
            col["number"] = {
                "min": float(min(nums)),
                "max": float(max(nums)),
                "avg": float(sum(nums) / max(1, len(nums))),
                "ok": int(num_ok),
                "bad": int(num_bad),
            }
        else:
            col["number"] = {"ok": int(num_ok), "bad": int(num_bad)}
        if dates:
            try:
                dates_sorted = sorted(dates)
                col["date"] = {
                    "min": dates_sorted[0].isoformat(),
                    "max": dates_sorted[-1].isoformat(),
                    "ok": int(date_ok),
                    "bad": int(date_bad),
                }
            except Exception:
                col["date"] = {"ok": int(date_ok), "bad": int(date_bad)}
        else:
            col["date"] = {"ok": int(date_ok), "bad": int(date_bad)}

        out["columns"][h] = col

    return out


def _group_rows_for_modules(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Rule-based grouping for mixed tables.
    Returns counts + masked representative samples per group.
    """
    # derive module mappings from headers
    map_act = _suggest_mapping_activities(headers)
    map_content = _suggest_mapping_content(headers)
    map_budget = _suggest_mapping_budget(headers)
    map_crm = _suggest_mapping_crm(headers)

    def cell(row: Dict[str, Any], header: Optional[str]) -> Any:
        if not header:
            return None
        return row.get(header)

    def has_any(row: Dict[str, Any], mapping: Dict[str, Optional[str]], fields: List[str]) -> int:
        c = 0
        for f in fields:
            if not _is_blankish(cell(row, mapping.get(f))):
                c += 1
        return c

    groups = {"activities": [], "content": [], "budget": [], "crm": [], "other": []}  # masked samples
    counts = {k: 0 for k in groups.keys()}

    for row in rows:
        # Budget
        b_score = 0
        if not _is_blankish(cell(row, map_budget.get("category"))) and not _is_blankish(cell(row, map_budget.get("amount"))):
            b_score += 3
        if not _is_blankish(cell(row, map_budget.get("metric"))) and not _is_blankish(cell(row, map_budget.get("target"))):
            b_score += 3
        if not _is_blankish(cell(row, map_budget.get("period"))):
            b_score += 1

        # Content
        c_score = 0
        if not _is_blankish(cell(row, map_content.get("title"))):
            c_score += 2
        c_score += has_any(row, map_content, ["channel", "format", "status", "due_at", "scheduled_at", "tags", "brief", "body"])

        # Activities
        a_score = 0
        if not _is_blankish(cell(row, map_act.get("title"))):
            a_score += 2
        a_score += has_any(row, map_act, ["category", "status", "budget", "start", "end", "notes", "weight"])

        # CRM
        crm_score = 0
        if not _is_blankish(cell(row, map_crm.get("company_name"))):
            crm_score += 3
        crm_score += has_any(row, map_crm, ["contact_name", "contact_email", "deal_title", "deal_value"])

        # Pick best
        scores = [
            ("budget", b_score),
            ("content", c_score),
            ("activities", a_score),
            ("crm", crm_score),
        ]
        best_kind, best = max(scores, key=lambda x: x[1])
        chosen = best_kind if best >= 4 else "other"

        counts[chosen] += 1
        if len(groups[chosen]) < 10:
            groups[chosen].append({h: _mask_value(row.get(h)) for h in headers[:80]})

    return {"counts": counts, "samples": groups, "mappings": {"activities": map_act, "content": map_content, "budget": map_budget, "crm": map_crm}}


def _anomaly_notes(headers: List[str], rows: List[Dict[str, Any]], stats: Dict[str, Any], group: Dict[str, Any]) -> List[str]:
    notes: List[str] = []
    n = int(stats.get("row_count") or len(rows) or 0)
    cols = stats.get("columns") or {}
    if not n:
        return notes

    # Many missing columns
    high_missing = [h for h, s in cols.items() if float(s.get("missing_ratio") or 0) >= 0.6]
    if len(high_missing) >= 6:
        notes.append(f"Viele Spalten sind stark leer (≥60% missing): z.B. {', '.join(high_missing[:6])}.")

    # Status normalization hint
    st_h = _suggest_header(headers, "status", "phase", "workflow")
    if st_h:
        vals = [str(r.get(st_h) or "").strip() for r in rows if not _is_blankish(r.get(st_h))]
        uniq = set(v for v in vals if v)
        if len(uniq) >= 10:
            notes.append(f"Status hat sehr viele Varianten ({len(uniq)}). Empfehlung: Normalisieren (z.B. geplant/aktiv/done/blocked).")

    # Date parse issues
    date_bad_cols = []
    for h, s in cols.items():
        if s.get("inferred_type") == "date":
            bad = int((s.get("date") or {}).get("bad") or 0)
            if bad >= max(3, int(0.05 * n)):
                date_bad_cols.append(h)
    if date_bad_cols:
        notes.append(f"Datumsfelder mit vielen unlesbaren Werten: {', '.join(date_bad_cols[:5])}.")

    # Number parse issues
    num_bad_cols = []
    for h, s in cols.items():
        if s.get("inferred_type") == "number":
            bad = int((s.get("number") or {}).get("bad") or 0)
            if bad >= max(3, int(0.05 * n)):
                num_bad_cols.append(h)
    if num_bad_cols:
        notes.append(f"Zahlenfelder mit vielen unlesbaren Werten: {', '.join(num_bad_cols[:5])} (z.B. Tausendertrennzeichen/Währung).")

    # Mixed-table detection
    counts = (group.get("counts") or {}) if isinstance(group, dict) else {}
    if sum(int(v or 0) for v in counts.values()) > 0:
        top = sorted([(k, int(v or 0)) for k, v in counts.items()], key=lambda x: x[1], reverse=True)
        if len([1 for _k, v in top if v > 0]) >= 3:
            notes.append("Tabelle wirkt gemischt (mehrere Module gleichzeitig). Smart Import ist sinnvoll; alternativ: separate Sheets/Tabellenbereiche.")

    return notes

@router.post("/ai-analyze")
async def ai_analyze_upload(
    request: Request,
    file: UploadFile = File(...),
    import_kind: Optional[str] = Form(default="activities"),
    current_user: User = Depends(get_current_user),
):
    """
    Analyze a CSV/XLSX and suggest import mapping + clean rules + high-level insights.
    Privacy: we only send masked sample rows (not the full file) to OpenAI.
    """
    enforce_rate_limit(request, scope="uploads_ai_analyze", limit=12, window_seconds=60)

    settings = get_settings()
    content = file.file.read()
    ctype = (file.content_type or "").lower()
    name_lower = (file.filename or "").lower()

    # Parse full file (best-effort) — we analyze ALL rows locally.
    headers: List[str] = []
    parsed_rows: List[Dict[str, Any]] = []
    sheet_name: Optional[str] = None
    tables_meta: List[Dict[str, Any]] = []
    if ctype in {"text/csv", "application/csv", "text/plain"} or name_lower.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            parsed_rows.append(row)
            if i >= 4999:
                break
        tables_meta = [{"sheet": "CSV", "rows": len(parsed_rows), "cols": len(headers)}]
    elif name_lower.endswith(".xlsx") or ctype in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
        tables = _extract_xlsx_tables(content)
        tables_meta = [{"sheet": t.get("sheet"), "rows": len(t.get("rows") or []), "cols": len(t.get("headers") or [])} for t in tables]
        # Heuristic: pick the largest table's headers, but analyze ALL rows across sheets.
        best = max(tables, key=lambda t: len(t.get("rows") or []))
        sheet_name = str(best.get("sheet") or "")
        headers = list(best.get("headers") or [])
        # Merge rows across all sheets; keep a soft cap
        merged: List[Dict[str, Any]] = []
        for t in tables:
            for r in list(t.get("rows") or []):
                merged.append(r)
                if len(merged) >= 8000:
                    break
            if len(merged) >= 8000:
                break
        parsed_rows = merged
    else:
        raise HTTPException(status_code=415, detail="Unsupported file type. Upload CSV or XLSX")

    kind = (import_kind or "activities").strip().lower()
    if kind not in {"activities", "crm", "content", "budget"}:
        kind = "activities"

    # Baseline deterministic suggestions
    if kind == "crm":
        base_mapping = _suggest_mapping_crm(headers)
    elif kind == "content":
        base_mapping = _suggest_mapping_content(headers)
    elif kind == "budget":
        base_mapping = _suggest_mapping_budget(headers)
    else:
        # reuse preview's local heuristic
        def _s(headers_in: List[str]) -> Dict[str, Optional[str]]:
            headers_l = [h.lower().replace(" ", "_") for h in headers_in]

            def suggest(*names: str) -> Optional[str]:
                for n in names:
                    if n in headers_l:
                        return headers_in[headers_l.index(n)]
                return None

            return {
                "title": suggest("title", "name", "massnahme", "maßnahme", "aktion", "initiative", "kampagne", "projekt"),
                "category": suggest(*_ACTIVITY_CATEGORY_HEADER_ALIASES),
                "status": suggest("status", "phase", "workflow"),
                "budget": suggest("budget", "budgetchf", "kosten", "betrag", "chf"),
                "notes": suggest("notes", "expected_output", "beschreibung", "kommentar", "notiz", "bemerkung"),
                "start": suggest("start", "start_date", "beginn", "von", "ab", "startdatum"),
                "end": suggest("end", "end_date", "ende", "bis", "enddatum"),
                "weight": suggest("weight", "gewicht", "prio", "priority"),
            }

        base_mapping = _s(headers)

    # Full-table statistics + grouping (local, private)
    stats = _column_stats(headers, parsed_rows)
    group = _group_rows_for_modules(headers, parsed_rows)
    anomaly_notes = _anomaly_notes(headers, parsed_rows, stats, group)

    # Representative samples: take grouped samples (masked)
    masked_samples = []
    for k in ["budget", "content", "activities", "crm", "other"]:
        for r in (group.get("samples") or {}).get(k, [])[:6]:
            masked_samples.append({"__group": k, **(r or {})})
    sample_n = len(masked_samples)

    missingness = _compute_missingness(headers, parsed_rows[:min(400, len(parsed_rows))]) if parsed_rows else {}
    top_cats = []
    cat_h = _suggest_header(headers, *_ACTIVITY_CATEGORY_HEADER_ALIASES)
    if cat_h:
        top_cats = _top_values(parsed_rows[:min(600, len(parsed_rows))], cat_h, limit=6)

    # Fallback (no OpenAI): return deterministic mapping + simple insights
    if not settings.openai_api_key:
        return {
            "ok": True,
            "provider": "fallback",
            "kind": kind,
            "recommended_kinds": [{"kind": kind, "score": 0.6, "reason": "Regel-basiert (kein OPENAI_API_KEY konfiguriert)."}],
            "suggested_mapping": base_mapping,
            "confidence": {k: (0.7 if v else 0.0) for k, v in base_mapping.items()},
            "clean_rules": [
                "Leere Werte wie '—'/'n/a' ignorieren",
                "Zahlen: 1'234.50 und 1.234,50 unterstützen",
                "Datumsfelder: YYYY-MM-DD und DD.MM.YYYY unterstützen",
            ],
            "insights": {
                "rows_scanned": len(parsed_rows),
                "rows_sampled": sample_n,
                "tables": tables_meta,
                "group_counts": group.get("counts"),
                "missingness": {k: float(v) for k, v in list(missingness.items())[:25]},
                "top_categories": top_cats,
                "notes": (anomaly_notes or []) + ["AI ist optional. Für bessere Vorschläge: OPENAI_API_KEY setzen."],
                "column_stats": stats.get("columns"),
            },
        }

    # OpenAI: ask for mapping + clean rules + module recommendations
    system = (
        "Du bist ein Daten-Import Assistent für eine Marketing-CRM Plattform. "
        "Du bekommst Header, Volltabellen-Statistiken (column_stats), erkannte Gruppen (group_counts) "
        "und maskierte repräsentative Beispielzeilen (samples). "
        "Gib nur dichte, umsetzbare Empfehlungen (keine Floskeln).\n\n"
        "Deine Aufgabe:\n"
        "1) Empfehle Import-Modus (activities|crm|content|budget) mit Score+Begründung\n"
        "2) Schlage Mapping für current_kind vor (field->headerName oder null) + confidence je Feld\n"
        "3) Gib Clean-Regeln (max 12, konkret)\n"
        "4) Gib Insights als kurze bullets + konkrete To-dos (z.B. 'Status normalisieren: ...').\n\n"
        "Antworte ausschließlich als JSON mit exakt diesen Keys:\n"
        "{"
        "\"recommended_kinds\": [{\"kind\":\"activities|crm|content|budget\",\"score\":0..1,\"reason\":\"...\"}],"
        "\"suggested_mapping\": {\"field\": \"header\"|null},"
        "\"confidence\": {\"field\": 0..1},"
        "\"clean_rules\": [\"...\"],"
        "\"insights\": {"
        "\"notes\":[\"...\"],"
        "\"todo\":[{\"title\":\"...\",\"why\":\"...\",\"how\":\"...\"}],"
        "\"top_categories\":[{\"value\":\"...\",\"count\":n}],"
        "\"budget_range_chf\":{\"min\":n,\"max\":n},"
        "\"period_guess\":{\"from\":\"...\",\"to\":\"...\"}"
        "}"
        "}"
        "Kein Markdown, kein zusätzlicher Text."
    )

    user_msg = {
        "filename": file.filename,
        "sheet": sheet_name,
        "current_kind": kind,
        "headers": headers[:120],
        "samples": masked_samples,
        "baseline_mapping": base_mapping,
        "tables": tables_meta,
        "group_counts": group.get("counts"),
        "column_stats": stats.get("columns"),
        "anomaly_notes": anomaly_notes,
        "missingness": {k: float(v) for k, v in list(missingness.items())[:40]},
        "top_categories": top_cats,
    }

    payload = {
        "model": settings.openai_model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": json.dumps(user_msg, ensure_ascii=False)},
        ],
        "temperature": 0.2,
        "max_tokens": 900,
        "response_format": {"type": "json_object"},
    }
    api_key = (settings.openai_api_key or "").strip()
    if not api_key:
        return {
            "ok": True,
            "provider": "fallback",
            "kind": kind,
            "recommended_kinds": [{"kind": kind, "score": 0.55, "reason": "Fallback (OPENAI_API_KEY fehlt)."}],
            "suggested_mapping": base_mapping,
            "confidence": {k: (0.7 if v else 0.0) for k, v in base_mapping.items()},
            "clean_rules": [
                "Leere Werte wie '—'/'n/a' ignorieren",
                "Zahlen: 1'234.50 und 1.234,50 unterstützen",
                "Datumsfelder: YYYY-MM-DD und DD.MM.YYYY unterstützen",
            ],
            "insights": {
                "rows_scanned": len(parsed_rows),
                "rows_sampled": sample_n,
                "tables": tables_meta,
                "group_counts": group.get("counts"),
                "missingness": {k: float(v) for k, v in list(missingness.items())[:25]},
                "top_categories": top_cats,
                "notes": (anomaly_notes or []) + ["OPENAI_API_KEY ist nicht gesetzt oder leer."],
                "column_stats": stats.get("columns"),
            },
        }

    headers_req = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    try:
        # Keep this comfortably below Vercel/Proxy timeouts to avoid request-level timeouts.
        async with httpx.AsyncClient(timeout=25.0) as client:
            r = await client.post("https://api.openai.com/v1/chat/completions", json=payload, headers=headers_req)
            if r.status_code == 400:
                # Some models may not support response_format. Retry once without it.
                payload2 = dict(payload)
                payload2.pop("response_format", None)
                r = await client.post("https://api.openai.com/v1/chat/completions", json=payload2, headers=headers_req)
            if r.status_code >= 400:
                # Avoid leaking sensitive details; include only status and a short hint.
                raise RuntimeError(f"openai_http_{r.status_code}")
            data = r.json()
            reply = data.get("choices", [{}])[0].get("message", {}).get("content", "") or ""
            try:
                obj = _safe_json_from_model_reply(reply)
            except Exception:
                raise RuntimeError("openai_invalid_json")

            suggested_mapping = obj.get("suggested_mapping") or {}
            confidence = obj.get("confidence") or {}
            clean_rules = obj.get("clean_rules") or []
            insights = obj.get("insights") or {}
            recommended_kinds = obj.get("recommended_kinds") or [{"kind": kind, "score": 0.5, "reason": "AI"}]

            merged_mapping = dict(base_mapping)
            for k2, v2 in dict(suggested_mapping).items():
                if k2 in merged_mapping:
                    merged_mapping[k2] = v2

            return {
                "ok": True,
                "provider": "openai",
                "kind": kind,
                "recommended_kinds": recommended_kinds,
                "suggested_mapping": merged_mapping,
                "confidence": confidence,
                "clean_rules": clean_rules,
                "insights": {
                    "rows_scanned": len(parsed_rows),
                    "rows_sampled": sample_n,
                    "missingness": {k: float(v) for k, v in list(missingness.items())[:25]},
                    **insights,
                },
            }
    except Exception as e:
        reason = "AI Analyse war nicht verfügbar, Mapping basiert auf Regeln."
        try:
            msg = str(e or "")
            if "openai_http_" in msg:
                code = msg.split("openai_http_")[-1].strip().split()[0]
                reason = f"OpenAI Anfrage fehlgeschlagen (HTTP {code}). Prüfe OPENAI_API_KEY/Model/Quota."
            elif "openai_invalid_json" in msg:
                reason = "OpenAI Antwort war kein gültiges JSON. Bitte erneut versuchen (oder Modell wechseln)."
            elif isinstance(e, json.JSONDecodeError):
                reason = "OpenAI Antwort konnte nicht als JSON gelesen werden (Decode-Fehler)."
            elif isinstance(e, httpx.TimeoutException):
                reason = "OpenAI Timeout. Bitte später erneut versuchen oder Timeout erhöhen."
            elif isinstance(e, httpx.ConnectError):
                reason = "OpenAI Connection Error (Netzwerk/DNS)."
            elif isinstance(e, httpx.HTTPError):
                reason = "OpenAI Request Error (Transport/Netzwerk)."
        except Exception:
            pass
        return {
            "ok": True,
            "provider": "fallback",
            "kind": kind,
            "recommended_kinds": [{"kind": kind, "score": 0.55, "reason": "Fallback (AI nicht erreichbar)."}],
            "suggested_mapping": base_mapping,
            "confidence": {k: (0.7 if v else 0.0) for k, v in base_mapping.items()},
            "clean_rules": [
                "Leere Werte wie '—'/'n/a' ignorieren",
                "Zahlen: 1'234.50 und 1.234,50 unterstützen",
                "Datumsfelder: YYYY-MM-DD und DD.MM.YYYY unterstützen",
            ],
            "insights": {
                "rows_scanned": len(parsed_rows),
                "rows_sampled": sample_n,
                "tables": tables_meta,
                "group_counts": group.get("counts"),
                "missingness": {k: float(v) for k, v in list(missingness.items())[:25]},
                "top_categories": top_cats,
                "notes": (anomaly_notes or []) + [reason, f"Diag: {type(e).__name__}"],
                "column_stats": stats.get("columns"),
            },
        }


@router.post("/smart-import")
def smart_import_upload(
    request: Request,
    file: Optional[UploadFile] = File(None),
    retry_upload_id: Optional[int] = Form(None),
    db: Session = Depends(get_db_session),
    current_user: User = Depends(require_writable_user),
):
    """
    Smart import for mixed tables.
    Use retry_upload_id to re-run import with same file (e.g. after fixing mapping).
    """
    enforce_rate_limit(request, scope="uploads_smart_import", limit=8, window_seconds=60)
    settings = get_settings()
    org = get_org_id(current_user)
    can_manage_all = current_user.role in {UserRole.admin, UserRole.editor}

    if retry_upload_id:
        upload = db.query(Upload).filter(Upload.id == retry_upload_id, Upload.organization_id == org).first()
        if not upload:
            raise HTTPException(status_code=404, detail="Upload not found")
        if not can_manage_all and int(getattr(upload, "owner_id", 0) or 0) != int(current_user.id):
            raise HTTPException(status_code=403, detail="Not allowed")
        content = bytes(getattr(upload, "content", None) or b"")
        if not content:
            raise HTTPException(status_code=400, detail="Upload content not stored; cannot retry")
        ctype = (getattr(upload, "file_type", "") or "").lower()
        name_lower = (getattr(upload, "original_name", "") or "").lower()
        # Cascade delete existing imported data before retry
        for model in [CalendarEntry, Activity, ContentTask, ContentItem, BudgetTarget, KpiTarget, Deal, Contact, Company, UserCategory]:
            try:
                db.query(model).filter(
                    model.organization_id == org,  # type: ignore[attr-defined]
                    model.source_upload_id == int(retry_upload_id),  # type: ignore[attr-defined]
                ).delete(synchronize_session=False)
            except Exception:
                pass
        db.commit()
        db.refresh(upload)
    else:
        if not file:
            raise HTTPException(status_code=400, detail="File or retry_upload_id required")
        content = file.file.read() or b""
        if len(content) > settings.upload_max_bytes:
            raise HTTPException(status_code=413, detail=f"File too large for current plan (max {settings.upload_max_bytes} bytes).")
        upload = Upload(
            original_name=file.filename or "file",
            file_type=file.content_type or "",
            file_size=len(content),
            organization_id=org,
            owner_id=current_user.id,
            sha256=hashlib.sha256(content).hexdigest(),
            stored_in_db=bool(settings.upload_store_in_db),
        )
        if settings.upload_store_in_db:
            upload.content = content
        db.add(upload)
        db.commit()
        db.refresh(upload)
        ctype = (file.content_type or "").lower()
        name_lower = (file.filename or "").lower()

    tables: List[Dict[str, Any]] = []
    if ctype in {"text/csv", "application/csv", "text/plain"} or name_lower.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 4999:
                break
        tables = [{"sheet": "CSV", "headers": headers, "rows": rows}]
    elif name_lower.endswith(".xlsx") or ctype in {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}:
        tables = _extract_xlsx_tables(content)
    else:
        raise HTTPException(status_code=415, detail="Unsupported file type. Upload CSV or XLSX")

    # Suggested mappings per module (per table)
    totals = {
        "activities_created": 0,
        "activities_skipped": 0,
        "content_created": 0,
        "content_skipped": 0,
        "tasks_created": 0,
        "tasks_skipped": 0,
        "calendar_created": 0,
        "budget_rows_applied": 0,
        "budget_rows_skipped": 0,
    }
    row_errors: List[Dict[str, Any]] = []

    # Create job early for progress visibility (phase: parsing -> analyzing -> importing -> upserting)
    total_rows = max(1, sum(len(t.get("rows") or []) for t in tables))
    job = Job(
        rq_id=f"local-smart-{upload.id}",
        type="import_smart",
        status="processing",
        phase="analyzing",
        progress=10,
        organization_id=org,
        upload_id=int(upload.id),
    )
    db.add(job)
    db.flush()

    default_period = _current_period()
    # Accumulate SAP Mediaplan KPI targets by quarter (period -> sums).
    # We upsert them once after all tables are processed.
    sap_kpi_acc: Dict[str, Dict[str, float]] = {}
    # Accumulate SAP Mediaplan budget targets by quarter (period+category -> amount).
    # We upsert them once after all tables are processed (overwrite for plan-correctness).
    sap_budget_acc: Dict[tuple[str, str], float] = {}

    def _is_sap_mediaplan_table(headers: List[str]) -> bool:
        hs = {str(h or "").strip().lower() for h in (headers or [])}
        return (
            "section" in hs
            and "market" in hs
            and "title" in hs
            and "start" in hs
            and "end" in hs
        )

    def _bucket_for_mediaplan_row(row: Dict[str, Any]) -> str:
        section = str(row.get("section") or "").strip()
        medium = str(row.get("medium") or "").strip()
        form = str(row.get("form") or "").strip()
        blob = f"{section} {medium} {form}".lower()
        if "sap intern" in (section or "").lower():
            return "SAP intern"
        if "paid social" in blob:
            return "Paid Social"
        if "organic" in blob:
            return "Organic Social"
        # Print often appears as "... Print" or printed version
        if " print" in blob or blob.endswith("print") or "magazin gedruckte" in blob:
            return "Print"
        return "Online Ads"

    def _activity_type_for_bucket(bucket: str) -> ActivityType:
        b = (bucket or "").strip().lower()
        if b == "print":
            return ActivityType.branding
        if b in {"paid social", "online ads"}:
            return ActivityType.sales
        if b == "organic social":
            return ActivityType.branding
        if b == "sap intern":
            return ActivityType.kundenpflege
        return ActivityType.sales

    def _budget_category_for_type(t: ActivityType) -> str:
        if t == ActivityType.branding:
            return "IMAGE"
        if t == ActivityType.employer_branding:
            return "EMPLOYER_BRANDING"
        if t == ActivityType.kundenpflege:
            return "KUNDENPFLEGE"
        return "VERKAUFSFOERDERUNG"

    def _period_from_dt(dt) -> str:
        try:
            d = dt.date() if hasattr(dt, "date") else dt
            y = int(getattr(d, "year", 0) or 0)
            m = int(getattr(d, "month", 0) or 0)
            if y and m:
                q = (m - 1) // 3 + 1
                return f"{y}-Q{q}"
        except Exception:
            pass
        return default_period

    def _overlap_days(a_start, a_end, b_start, b_end) -> int:
        try:
            s = max(a_start, b_start)
            e = min(a_end, b_end)
            if e < s:
                return 0
            return int((e - s).days) + 1
        except Exception:
            return 0

    def _quarter_bounds(y: int, q: int):
        m1 = (q - 1) * 3 + 1
        m2 = q * 3
        last = calendar.monthrange(int(y), int(m2))[1]
        from datetime import date

        return date(int(y), int(m1), 1), date(int(y), int(m2), int(last))

    def _period_fracs_for_range(start_dt, end_dt) -> List[tuple[str, float]]:
        """
        Return overlapping quarters with fractional weights for the given range (inclusive, by days).
        """
        try:
            from datetime import date

            s = start_dt.date() if hasattr(start_dt, "date") else start_dt
            e = end_dt.date() if hasattr(end_dt, "date") else end_dt
            if not isinstance(s, date) or not isinstance(e, date):
                return []
            if e < s:
                s, e = e, s
            total = int((e - s).days) + 1
            if total <= 0:
                return []
            out: List[tuple[str, float]] = []
            for y in range(int(s.year), int(e.year) + 1):
                for q in (1, 2, 3, 4):
                    qs, qe = _quarter_bounds(y, q)
                    od = _overlap_days(s, e, qs, qe)
                    if od > 0:
                        out.append((f"{y}-Q{q}", float(od) / float(total)))
            return out
        except Exception:
            return []

    def _alloc_year_budget(year: int, amount: float, start_dt, end_dt) -> List[tuple[str, float]]:
        """
        Allocate a year-specific budget across the quarters of that year that overlap the schedule range.
        """
        try:
            from datetime import date

            amt = float(amount or 0)
            if amt <= 0:
                return []
            s = start_dt.date() if hasattr(start_dt, "date") else start_dt
            e = end_dt.date() if hasattr(end_dt, "date") else end_dt
            if not isinstance(s, date) or not isinstance(e, date):
                return []
            if e < s:
                s, e = e, s
            y = int(year)
            ys = max(s, date(y, 1, 1))
            ye = min(e, date(y, 12, 31))
            if ye < ys:
                return []
            total = int((ye - ys).days) + 1
            if total <= 0:
                return []
            out: List[tuple[str, float]] = []
            for q in (1, 2, 3, 4):
                qs, qe = _quarter_bounds(y, q)
                od = _overlap_days(ys, ye, qs, qe)
                if od > 0:
                    out.append((f"{y}-Q{q}", (amt * float(od)) / float(total)))
            return out
        except Exception:
            return []

    # If the upload looks like a SAP Mediaplan, auto-create a small, harmonious set of ring categories
    # and reuse them for activity.category_name (circle).
    ring_category_color: Dict[str, str] = {}
    ring_category_names: List[str] = []
    sap_company_id: Optional[int] = None
    try:
        # Pick the first SAP table to compute buckets.
        sap_tables = [t for t in tables if _is_sap_mediaplan_table(list(t.get("headers") or []))]
        if sap_tables:
            sap_rows = list((sap_tables[0].get("rows") or []))
            # Create a minimal default CRM company for new orgs so Dashboard/CRM isn't empty.
            # Only do this when the org has no companies yet (safe, avoids polluting existing CRMs).
            try:
                has_any_company = db.query(Company).filter(Company.organization_id == org).first()
                if not has_any_company:
                    c = Company(
                        organization_id=org,
                        name="SAP",
                        industry=None,
                        website=None,
                        email=None,
                        phone=None,
                        status="active",
                        lead_source="sap_mediaplan_import",
                        notes="Auto-created from SAP Mediaplan import.",
                        source_upload_id=int(upload.id),
                    )
                    db.add(c)
                    db.flush()
                    sap_company_id = int(getattr(c, "id", 0) or 0) or None
            except Exception:
                sap_company_id = None
            # Candidate buckets in a stable order
            preferred = ["Online Ads", "Print", "Paid Social", "Organic Social", "SAP intern"]
            present = []
            for name in preferred:
                if any(_bucket_for_mediaplan_row(r) == name for r in sap_rows):
                    present.append(name)
            # Fallback: whatever appears, capped
            if not present:
                seen = []
                for r in sap_rows:
                    b = _bucket_for_mediaplan_row(r)
                    if b not in seen:
                        seen.append(b)
                present = seen[:5]
            ring_category_names = present[:5]

            # Load existing user categories for this user+org
            existing = (
                db.query(UserCategory)
                .filter(UserCategory.user_id == current_user.id, UserCategory.organization_id == org)
                .order_by(UserCategory.position.asc(), UserCategory.id.asc())
                .all()
            )
            used_colors = {str(c.color or "").strip() for c in existing if getattr(c, "color", None)}
            name_to_color = {str(c.name or "").strip(): str(c.color or "").strip() for c in existing}
            max_pos = max([int(getattr(c, "position", 0) or 0) for c in existing] + [-1])

            # Ensure categories exist (do not delete/replace user's existing ones).
            for idx, name in enumerate(ring_category_names):
                if name in name_to_color and name_to_color[name]:
                    ring_category_color[name] = name_to_color[name]
                    continue
                color = None
                for cand in _CATEGORY_COLOR_PALETTE:
                    if cand not in used_colors:
                        color = cand
                        break
                if not color:
                    color = _CATEGORY_COLOR_PALETTE[(idx + max_pos + 1) % len(_CATEGORY_COLOR_PALETTE)]
                used_colors.add(color)
                item = UserCategory(
                    user_id=current_user.id,
                    organization_id=org,
                    name=name,
                    color=color,
                    position=max_pos + 1 + idx,
                    source_upload_id=int(upload.id),
                )
                db.add(item)
                ring_category_color[name] = color
            # Don't commit yet; commit happens at the end of smart import.
    except Exception:
        ring_category_color = {}
        ring_category_names = []

    table_row_offsets: List[int] = [0]
    for t in tables:
        table_row_offsets.append(table_row_offsets[-1] + len(t.get("rows") or []))
    for t_idx, t in enumerate(tables):
        headers = list(t.get("headers") or [])
        rows = list(t.get("rows") or [])
        if not headers or not rows:
            continue
        table_start = table_row_offsets[t_idx]

        map_act = _suggest_mapping_activities(headers)
        map_content = _suggest_mapping_content(headers)
        map_budget = _suggest_mapping_budget(headers)

        # Extra heuristics for Task Board imports (separate from ContentItem mapping)
        task_title_h = _suggest_header(headers, "task", "aufgabe", "todo", "massnahme", "massnahmen", "maßnahme", "maßnahmen", "aktion", "activity")
        task_deadline_h = _suggest_header(headers, "deadline", "due", "due_date", "faellig", "fällig", "stichtag", "abgabe", "abgabedatum", "end", "end_date", "ende", "enddatum", "bis")
        task_status_h = _suggest_header(headers, "task_status", "status", "phase", "workflow")
        task_priority_h = _suggest_header(headers, "priority", "prio", "dringend", "importance")
        item_title_h = _suggest_header(headers, "title", "titel", "thema", "subject", "betreff", "name")

        def cell(row: Dict[str, Any], header: Optional[str]) -> Any:
            if not header:
                return None
            return row.get(header)

        def _add_row_err(r_idx: int, msg: str, src: str, r: Dict[str, Any]) -> None:
            raw_preview = {str(k)[:50]: str(v)[:100] for k, v in list(r.items())[:8]}
            row_errors.append({"row": r_idx + 1, "message": str(msg)[:500], "source": src, "raw": raw_preview})

        job.phase = "importing"
        job.progress = 15
        db.flush()
        for row_idx, row in enumerate(rows):
            # --- Budget rows ---
            try:
                per_raw = cell(row, map_budget.get("period")) or default_period
                period = str(per_raw).strip() if per_raw not in (None, "") else default_period
                if not period:
                    period = default_period
                cat_raw = cell(row, map_budget.get("category"))
                amt_raw = cell(row, map_budget.get("amount"))
                metric_raw = cell(row, map_budget.get("metric"))
                target_raw = cell(row, map_budget.get("target"))
                unit_raw = cell(row, map_budget.get("unit"))

                did_budget = False
                amount = _parse_float(amt_raw)
                if cat_raw not in (None, "") and amount is not None:
                    category = _norm_budget_category(cat_raw)
                    existing = (
                        db.query(BudgetTarget)
                        .filter(BudgetTarget.organization_id == org, BudgetTarget.period == period, BudgetTarget.category == category)
                        .first()
                    )
                    if existing:
                        existing.amount_chf = float(amount)
                        existing.source_upload_id = int(upload.id)
                        db.add(existing)
                    else:
                        db.add(BudgetTarget(organization_id=org, period=period, category=category, amount_chf=float(amount), source_upload_id=int(upload.id)))
                    did_budget = True

                target_val = _parse_float(target_raw)
                if metric_raw not in (None, "") and target_val is not None:
                    m = str(metric_raw).strip()
                    if m:
                        existing_k = (
                            db.query(KpiTarget)
                            .filter(KpiTarget.organization_id == org, KpiTarget.period == period, KpiTarget.metric == m)
                            .first()
                        )
                        unit = str(unit_raw).strip()[:20] if unit_raw not in (None, "") else None
                        if existing_k:
                            existing_k.target_value = float(target_val)
                            if unit:
                                existing_k.unit = unit
                            existing_k.source_upload_id = int(upload.id)
                            db.add(existing_k)
                        else:
                            db.add(KpiTarget(organization_id=org, period=period, metric=m, target_value=float(target_val), unit=unit, source_upload_id=int(upload.id)))
                        did_budget = True

                if did_budget:
                    totals["budget_rows_applied"] += 1
                else:
                    totals["budget_rows_skipped"] += 1
            except Exception as e:
                totals["budget_rows_skipped"] += 1
                _add_row_err(row_idx, str(e), "budget", row)

            # --- Content rows ---
            try:
                c_title = cell(row, map_content.get("title"))
                c_channel = cell(row, map_content.get("channel"))
                c_scheduled = cell(row, map_content.get("scheduled_at"))
                c_due = cell(row, map_content.get("due_at"))
                c_body = cell(row, map_content.get("body"))
                c_brief = cell(row, map_content.get("brief"))

                # create only if looks like an editorial row (title + one of these)
                if not _is_blankish(c_title) and (
                    (not _is_blankish(c_channel))
                    or (not _is_blankish(c_scheduled))
                    or (not _is_blankish(c_due))
                    or (not _is_blankish(c_body))
                    or (not _is_blankish(c_brief))
                ):
                    # Prefer a "real" item title column if present (separate from task/action column).
                    title_raw = cell(row, item_title_h) if item_title_h else c_title
                    title = str(title_raw).strip() if not _is_blankish(title_raw) else str(c_title).strip()
                    channel = str(c_channel or "Website").strip()[:100]
                    fmt = cell(row, map_content.get("format"))
                    fmt_s = str(fmt).strip()[:100] if not _is_blankish(fmt) else None

                    raw_status = str(cell(row, map_content.get("status")) or "").strip().lower()
                    status_map = {
                        "idea": ContentItemStatus.IDEA,
                        "idee": ContentItemStatus.IDEA,
                        "draft": ContentItemStatus.DRAFT,
                        "entwurf": ContentItemStatus.DRAFT,
                        "review": ContentItemStatus.REVIEW,
                        "prüfung": ContentItemStatus.REVIEW,
                        "pruefung": ContentItemStatus.REVIEW,
                        "approved": ContentItemStatus.APPROVED,
                        "freigegeben": ContentItemStatus.APPROVED,
                        "scheduled": ContentItemStatus.SCHEDULED,
                        "geplant": ContentItemStatus.SCHEDULED,
                        "published": ContentItemStatus.PUBLISHED,
                        "veröffentlicht": ContentItemStatus.PUBLISHED,
                        "veroeffentlicht": ContentItemStatus.PUBLISHED,
                        "archived": ContentItemStatus.ARCHIVED,
                        "archiviert": ContentItemStatus.ARCHIVED,
                    }
                    status = status_map.get(raw_status, ContentItemStatus.DRAFT)

                    due_at = _parse_datetime_loose(cell(row, map_content.get("due_at")))
                    scheduled_at = _parse_datetime_loose(cell(row, map_content.get("scheduled_at")))
                    tags = _parse_tags(cell(row, map_content.get("tags")))
                    brief = cell(row, map_content.get("brief"))
                    body = cell(row, map_content.get("body"))
                    language = cell(row, map_content.get("language")) or "de"
                    tone = cell(row, map_content.get("tone"))

                    owner_id = current_user.id
                    owner_email = cell(row, map_content.get("owner_email"))
                    if not _is_blankish(owner_email):
                        try:
                            oe = str(owner_email).strip().lower()
                            u = db.query(User).filter(User.organization_id == org, User.email == oe).first()
                            if u:
                                owner_id = u.id
                        except Exception:
                            pass

                    item = ContentItem(
                        organization_id=org,
                        owner_id=owner_id,
                        source_upload_id=int(upload.id),
                        title=title[:255],
                        channel=channel,
                        format=fmt_s,
                        status=status,
                        tags=tags,
                        brief=str(brief).strip() if not _is_blankish(brief) else None,
                        body=str(body).strip() if not _is_blankish(body) else None,
                        language=str(language).strip()[:10] if not _is_blankish(language) else "de",
                        tone=str(tone).strip()[:50] if not _is_blankish(tone) else None,
                        due_at=due_at,
                        scheduled_at=scheduled_at,
                    )
                    db.add(item)
                    db.flush()
                    totals["content_created"] += 1

                    # --- Task Board: create ContentTask when row looks task-like ---
                    try:
                        # Prefer a dedicated task title column; otherwise fall back to item title.
                        t_title_raw = cell(row, task_title_h) if task_title_h else None
                        t_title = None
                        if not _is_blankish(t_title_raw):
                            t_title = str(t_title_raw).strip()
                        elif not _is_blankish(title):
                            # If no explicit task column exists, still create a task for the item
                            # when there is a deadline or workflow-ish status.
                            t_title = str(title).strip()

                        # Deadline
                        deadline = _parse_datetime_loose(cell(row, task_deadline_h)) if task_deadline_h else None
                        if deadline is None:
                            deadline = due_at

                        # Status
                        raw_t_status = str(cell(row, task_status_h) or raw_status or "").strip().lower()
                        t_status_map = {
                            "todo": ContentTaskStatus.TODO,
                            "to_do": ContentTaskStatus.TODO,
                            "open": ContentTaskStatus.TODO,
                            "offen": ContentTaskStatus.TODO,
                            "in_progress": ContentTaskStatus.IN_PROGRESS,
                            "in bearbeitung": ContentTaskStatus.IN_PROGRESS,
                            "bearbeitung": ContentTaskStatus.IN_PROGRESS,
                            "doing": ContentTaskStatus.IN_PROGRESS,
                            "review": ContentTaskStatus.REVIEW,
                            "prüfung": ContentTaskStatus.REVIEW,
                            "pruefung": ContentTaskStatus.REVIEW,
                            "approved": ContentTaskStatus.APPROVED,
                            "freigegeben": ContentTaskStatus.APPROVED,
                            "published": ContentTaskStatus.PUBLISHED,
                            "veröffentlicht": ContentTaskStatus.PUBLISHED,
                            "veroeffentlicht": ContentTaskStatus.PUBLISHED,
                            "archived": ContentTaskStatus.ARCHIVED,
                            "archiviert": ContentTaskStatus.ARCHIVED,
                            "done": ContentTaskStatus.PUBLISHED,
                            "completed": ContentTaskStatus.PUBLISHED,
                            "erledigt": ContentTaskStatus.PUBLISHED,
                        }
                        t_status = t_status_map.get(raw_t_status, ContentTaskStatus.TODO)

                        # Priority
                        raw_prio = str(cell(row, task_priority_h) or "").strip().lower() if task_priority_h else ""
                        prio_map = {
                            "low": ContentTaskPriority.LOW,
                            "niedrig": ContentTaskPriority.LOW,
                            "medium": ContentTaskPriority.MEDIUM,
                            "mittel": ContentTaskPriority.MEDIUM,
                            "high": ContentTaskPriority.HIGH,
                            "hoch": ContentTaskPriority.HIGH,
                            "urgent": ContentTaskPriority.URGENT,
                            "dringend": ContentTaskPriority.URGENT,
                        }
                        t_prio = prio_map.get(raw_prio, ContentTaskPriority.MEDIUM)

                        # Only create if we have at least a title AND some task signal (deadline or explicit task column)
                        has_explicit_task = task_title_h is not None and not _is_blankish(t_title_raw)
                        if t_title and (has_explicit_task or deadline is not None):
                            t_notes = None
                            if not _is_blankish(c_brief):
                                t_notes = str(c_brief).strip()[:2000]
                            elif not _is_blankish(c_body):
                                t_notes = str(c_body).strip()[:2000]
                            db.add(
                                ContentTask(
                                    organization_id=org,
                                    owner_id=owner_id,
                                    source_upload_id=int(upload.id),
                                    title=str(t_title).strip()[:255],
                                    channel=channel,
                                    format=fmt_s,
                                    status=t_status,
                                    priority=t_prio,
                                    notes=t_notes,
                                    deadline=deadline,
                                    content_item_id=int(item.id) if getattr(item, "id", None) is not None else None,
                                )
                            )
                            totals["tasks_created"] += 1
                        else:
                            totals["tasks_skipped"] += 1
                    except Exception as e2:
                        totals["tasks_skipped"] += 1
                        _add_row_err(row_idx, str(e2), "task", row)

                    # --- Calendar sync: mirror scheduled_at into CalendarEntry (global calendar) ---
                    try:
                        if scheduled_at:
                            from datetime import timedelta

                            ev = CalendarEntry(
                                title=f"Content: {item.title}",
                                description=(item.brief or None),
                                start_time=scheduled_at,
                                end_time=scheduled_at + timedelta(minutes=30),
                                event_type="content",
                                status="PLANNED",
                                category=(item.channel or "Content").strip() if item.channel else "Content",
                                priority="medium",
                                color="#a78bfa",
                                content_item_id=item.id,
                                source_upload_id=int(upload.id),
                                owner_id=owner_id,
                                organization_id=org,
                            )
                            db.add(ev)
                            totals["calendar_created"] += 1
                    except Exception:
                        pass
                else:
                    totals["content_skipped"] += 1
            except Exception as e:
                totals["content_skipped"] += 1
                _add_row_err(row_idx, str(e), "content", row)

            # --- Activities rows ---
            try:
                a_title = cell(row, map_act.get("title"))
                if _is_blankish(a_title):
                    totals["activities_skipped"] += 1
                else:
                    # create only if looks like an activity row (has at least category/status/budget/date)
                    cat = cell(row, map_act.get("category"))
                    bud = cell(row, map_act.get("budget"))
                    st = cell(row, map_act.get("status"))
                    sd = cell(row, map_act.get("start"))
                    ed = cell(row, map_act.get("end"))
                    if (
                        not _is_blankish(cat)
                        or not _is_blankish(bud)
                        or not _is_blankish(st)
                        or not _is_blankish(sd)
                        or not _is_blankish(ed)
                    ):
                        title = str(a_title).strip() or "Untitled"
                        is_sap = _is_sap_mediaplan_table(headers)
                        bucket = _bucket_for_mediaplan_row(row) if is_sap else None
                        category = str(cat or "VERKAUFSFOERDERUNG").strip()
                        if is_sap and bucket:
                            # Use ring bucket as the human category name, but keep Activity.type consistent.
                            category = bucket
                        status_raw = str(st or "ACTIVE").strip()
                        budget = _parse_float(bud)
                        weight = _parse_float(cell(row, map_act.get("weight")))
                        notes = cell(row, map_act.get("notes"))
                        start = _parse_datetime_loose(sd)
                        end = _parse_datetime_loose(ed)

                        a_type = _activity_type_for_bucket(category) if is_sap and category in ring_category_names else _map_category_to_activity_type(category)
                        act = Activity(
                            title=title,
                            type=a_type,
                            category_name=category,
                            status=status_raw.upper(),
                            source_upload_id=int(upload.id),
                            budget=budget,
                            weight=weight,
                            expected_output=str(notes).strip() if not _is_blankish(notes) else None,
                            start_date=start.date() if hasattr(start, "date") and start else None,
                            end_date=end.date() if hasattr(end, "date") and end else None,
                            owner_id=current_user.id,
                            organization_id=org,
                        )
                        db.add(act)
                        # Calendar entry for activities (helps make the import visible everywhere).
                        try:
                            if start:
                                from datetime import timedelta

                                st_dt = start
                                en_dt = end or start
                                # If end date is only a day, end event at +60min; if multi-day, end at end date + 1h.
                                if en_dt == st_dt:
                                    en_dt = st_dt + timedelta(minutes=60)
                                else:
                                    en_dt = en_dt + timedelta(minutes=60)
                                color = ring_category_color.get(category) if is_sap else None
                                ev = CalendarEntry(
                                    title=f"Activity: {title}",
                                    description=str(notes).strip() if not _is_blankish(notes) else None,
                                    start_time=st_dt,
                                    end_time=en_dt,
                                    event_type="activity",
                                    status=status_raw.upper(),
                                    category=category,
                                    priority="medium",
                                    color=color,
                                    company_id=(sap_company_id if (is_sap and sap_company_id) else None),
                                    source_upload_id=int(upload.id),
                                    activity=act,
                                    owner_id=current_user.id,
                                    organization_id=org,
                                )
                                db.add(ev)
                                totals["calendar_created"] += 1
                        except Exception:
                            pass

                        # Budget targets (aggregate planned spend into Budget module by quarter + high-level category).
                        try:
                            if budget is not None and start:
                                bcat = _budget_category_for_type(a_type)
                                if is_sap:
                                    # SAP Mediaplan: budgets are usually split per year (e.g. CHF_2022 / CHF_2023).
                                    # Allocate each year's budget across overlapping quarters of that year, then upsert once (overwrite).
                                    end_dt = end or start
                                    ybuds: Dict[int, float] = {}
                                    for ykey in ("2022", "2023"):
                                        v = _parse_float(row.get(f"chf_{ykey}"))
                                        if v is not None and float(v) > 0:
                                            ybuds[int(ykey)] = float(v)
                                    # Fallback: if year columns are absent, treat the combined budget as belonging to the start year.
                                    if not ybuds:
                                        try:
                                            y0 = int(getattr(start.date() if hasattr(start, "date") else start, "year", 0) or 0)
                                        except Exception:
                                            y0 = 0
                                        if y0 and float(budget) > 0:
                                            ybuds[y0] = float(budget)

                                    for y, yamt in ybuds.items():
                                        for per, amt in _alloc_year_budget(int(y), float(yamt), start, end_dt):
                                            key = (per, bcat)
                                            sap_budget_acc[key] = float(sap_budget_acc.get(key, 0.0) or 0.0) + float(amt or 0)
                                            totals["budget_rows_applied"] += 1
                                else:
                                    period = _period_from_dt(start)
                                    existing_b = (
                                        db.query(BudgetTarget)
                                        .filter(BudgetTarget.organization_id == org, BudgetTarget.period == period, BudgetTarget.category == bcat)
                                        .first()
                                    )
                                    if existing_b:
                                        existing_b.amount_chf = float(existing_b.amount_chf or 0) + float(budget)
                                        # Tag updates too so deletion can fully revert an import.
                                        existing_b.source_upload_id = int(upload.id)
                                        db.add(existing_b)
                                    else:
                                        db.add(
                                            BudgetTarget(
                                                organization_id=org,
                                                period=period,
                                                category=bcat,
                                                amount_chf=float(budget),
                                                source_upload_id=int(upload.id),
                                            )
                                        )
                                    totals["budget_rows_applied"] += 1
                        except Exception:
                            pass

                        # KPI targets from SAP Mediaplan (impressions/clicks/views/CTR), aggregated per quarter.
                        try:
                            if is_sap and start:
                                imp = _parse_float(row.get("impressions"))
                                clk = _parse_float(row.get("clicks"))
                                vw = _parse_float(row.get("views_adobe"))
                                ctrp = _parse_percent(row.get("ctr"))
                                end_dt = end or start
                                fracs = _period_fracs_for_range(start, end_dt)
                                for period_kpi, frac in fracs:
                                    acc = sap_kpi_acc.setdefault(
                                        period_kpi,
                                        {"impressions": 0.0, "clicks": 0.0, "views": 0.0, "ctr_wsum": 0.0, "ctr_w": 0.0},
                                    )
                                    f = float(frac or 0.0)
                                    if f <= 0:
                                        continue
                                    if imp is not None:
                                        acc["impressions"] += float(imp) * f
                                    if clk is not None:
                                        acc["clicks"] += float(clk) * f
                                    if vw is not None:
                                        acc["views"] += float(vw) * f
                                    if ctrp is not None:
                                        w = float(imp or clk or 1.0) * f
                                        acc["ctr_wsum"] += float(ctrp) * w
                                        acc["ctr_w"] += w
                        except Exception:
                            pass

                        totals["activities_created"] += 1
                    else:
                        totals["activities_skipped"] += 1
            except Exception as e:
                totals["activities_skipped"] += 1
                _add_row_err(row_idx, str(e), "activities", row)

            # Chunked commit: every 500 rows to avoid timeouts on large XLSX
            if (row_idx + 1) % 500 == 0:
                db.commit()
                rows_done = table_start + row_idx + 1
                job.progress = min(85, 15 + 70 * rows_done // total_rows)
                db.flush()

        db.commit()

    job.phase = "upserting"
    job.progress = 90
    db.flush()
    # Upsert derived KPI targets from SAP Mediaplan tables (best-effort).
    # This makes KPI targets visible on Budget/Performance pages even for historic plans.
    try:
        # Upsert SAP-derived budget targets (overwrite per period/category for plan-correctness).
        for (period, category), amount in (sap_budget_acc or {}).items():
            try:
                p = str(period or "").strip()
                c = str(category or "").strip().upper()
                amt = float(amount or 0)
                if not p or not c or amt <= 0:
                    continue
                existing_b = (
                    db.query(BudgetTarget)
                    .filter(BudgetTarget.organization_id == org, BudgetTarget.period == p, BudgetTarget.category == c)
                    .first()
                )
                if existing_b:
                    existing_b.amount_chf = float(amt)
                    existing_b.source_upload_id = int(upload.id)
                    db.add(existing_b)
                else:
                    db.add(BudgetTarget(organization_id=org, period=p, category=c, amount_chf=float(amt), source_upload_id=int(upload.id)))
            except Exception:
                continue

        for period, acc in (sap_kpi_acc or {}).items():
            if not period:
                continue

            def _upsert(metric: str, value: float, unit: Optional[str]) -> None:
                existing_k = (
                    db.query(KpiTarget)
                    .filter(KpiTarget.organization_id == org, KpiTarget.period == period, KpiTarget.metric == metric)
                    .first()
                )
                if existing_k:
                    existing_k.target_value = float(value or 0)
                    if unit:
                        existing_k.unit = unit
                    existing_k.source_upload_id = int(upload.id)
                    db.add(existing_k)
                else:
                    db.add(
                        KpiTarget(
                            organization_id=org,
                            period=period,
                            metric=metric,
                            target_value=float(value or 0),
                            unit=unit,
                            source_upload_id=int(upload.id),
                        )
                    )

            imps = float(acc.get("impressions") or 0)
            clicks = float(acc.get("clicks") or 0)
            views = float(acc.get("views") or 0)
            ctr_w = float(acc.get("ctr_w") or 0)
            ctr_wsum = float(acc.get("ctr_wsum") or 0)

            if imps > 0:
                _upsert("Impressions", imps, "Impr.")
            if clicks > 0:
                _upsert("Clicks", clicks, "Clicks")
            if views > 0:
                _upsert("Views", views, "Views")
            if ctr_w > 0:
                _upsert("CTR", ctr_wsum / ctr_w, "%")
    except Exception:
        pass

    # Update job with final result (job was created at start)
    job_result = {**totals}
    if row_errors:
        job_result["row_errors"] = row_errors
    job.status = "finished"
    job.result = json.dumps(job_result)
    job.phase = "finished"
    job.progress = 100
    db.flush()
    # Audit log: who imported
    if settings.upload_audit_enabled:
        _safe_add_upload_audit_log(
            db,
            organization_id=org,
            upload_id=int(upload.id),
            actor_id=current_user.id,
            action="imported",
            details=json.dumps({"job_id": int(job.id), "totals": totals}),
        )
    db.commit()

    return {
        "ok": True,
        "upload_id": int(upload.id),
        "job_id": int(job.id),
        "import": totals,
        "row_errors_count": len(row_errors),
        "tables": [{"sheet": t.get("sheet"), "rows": len(t.get("rows") or []), "cols": len(t.get("headers") or [])} for t in tables],
    }


@router.post("/cron/retention")
def cron_retention(
    token: Optional[str] = Query(None, alias="token"),
    db: Session = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    Auto-delete uploads older than UPLOAD_RETENTION_DAYS.
    Call via cron (e.g. Render Cron Job) with UPLOAD_RETENTION_CRON_TOKEN.
    """
    settings = get_settings()
    if settings.upload_retention_days <= 0:
        return {"ok": True, "deleted": 0, "reason": "retention disabled"}
    if not settings.upload_retention_cron_token or token != settings.upload_retention_cron_token:
        raise HTTPException(status_code=403, detail="Invalid or missing token")
    cutoff = datetime.now(timezone.utc) - timedelta(days=settings.upload_retention_days)
    uploads = db.query(Upload).filter(Upload.created_at < cutoff).limit(500).all()
    deleted_count = 0
    for u in uploads:
        try:
            org = getattr(u, "organization_id", None)
            if org:
                for model in [
                    CalendarEntry, Activity, ContentTask, ContentItem,
                    BudgetTarget, KpiTarget, Deal, Contact, Company, UserCategory,
                ]:
                    try:
                        db.query(model).filter(
                            model.organization_id == org,  # type: ignore[attr-defined]
                            model.source_upload_id == int(u.id),  # type: ignore[attr-defined]
                        ).delete(synchronize_session=False)
                    except Exception:
                        pass
            db.delete(u)
            deleted_count += 1
        except Exception:
            db.rollback()
            continue
    try:
        db.commit()
    except Exception:
        db.rollback()
    return {"ok": True, "deleted": deleted_count, "cutoff": cutoff.isoformat()}


@router.get("/template.csv")
def template_csv(
    current_user: User = Depends(get_current_user),
) -> Response:
    headers = ["Title","Category","Status","BudgetCHF","Start","End","Notes","Weight"]
    example1 = ["Sommer Aktion","VERKAUFSFOERDERUNG","ACTIVE","1200","2025-06-01","2025-06-30","Promo Juni","1"]
    example2 = ["Brand Kampagne","IMAGE","PLANNED","5000","01.07.2025","31.07.2025","Awareness","2"]
    lines = [",".join(headers), ",".join(map(str, example1)), ",".join(map(str, example2))]
    content = "\n".join(lines)
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=activities-template.csv"},
    )


@router.get("/template.xlsx")
def template_xlsx(
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    try:
        import openpyxl  # type: ignore
    except Exception:
        # Fallback: serve CSV if xlsx unsupported
        return StreamingResponse(
            iter([b"Title,Category,Status,BudgetCHF,Start,End,Notes,Weight\nSommer Aktion,VERKAUFSFOERDERUNG,ACTIVE,1200,2025-06-01,2025-06-30,Promo Juni,1\nBrand Kampagne,IMAGE,PLANNED,5000,01.07.2025,31.07.2025,Awareness,2\n"]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=activities-template.csv"},
        )

    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Title","Category","Status","BudgetCHF","Start","End","Notes","Weight"])
    ws.append(["Sommer Aktion","VERKAUFSFOERDERUNG","ACTIVE",1200,"2025-06-01","2025-06-30","Promo Juni",1])
    ws.append(["Brand Kampagne","IMAGE","PLANNED",5000,"01.07.2025","31.07.2025","Awareness",2])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=activities-template.xlsx"},
    )


